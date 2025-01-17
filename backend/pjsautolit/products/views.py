import ast
import base64
import datetime
import io
import json
import logging
import os
import queue
import re
import threading
import time
import traceback
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from signal import signal
from threading import Lock

import requests
from bs4 import BeautifulSoup
from celery import shared_task
from django.conf import settings
from django.contrib import messages
from django.core.cache import cache
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import BigIntegerField, F, Q, Value
from django.http import HttpResponse, HttpResponseNotFound, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template import Context, Template
from django.utils import timezone
from django.views.decorators.http import require_GET
from django.views.generic.detail import DetailView
from requests.exceptions import HTTPError, RequestException

from .models import (Cart, CartItem, FetchStatus, Order, Product,
                     ProductChangeLog)

EBAY_APP_ID = settings.EBAY_APP_ID
EBAY_AUTH_TOKEN = settings.EBAY_AUTH_TOKEN
FINDING_API_URL = (
    "https://svcs.ebay.com/services/search/FindingService/v1"  # gets basic info
)
BROWSE_API_URL = (
    "https://api.ebay.com/buy/browse/v1/item/v1|{item_id}|0"  # gets detailed info
)

MAX_RETRIES = 5
RETRY_DELAY = 2  # Initial delay in seconds


def process_queue(q):
    while True:
        item = q.get()
        if item is None:
            break
        try:
            save_product_data(item)
        except Exception as e:
            print(f"Error saving product data: {e}")
        time.sleep(2)  # 2-second delay
        q.task_done()


@require_GET
def fetch_all_items(request):
    cache.set("fetch_items_progress", {"progress": 0, "completed": False})
    total_items = 0
    min_price = 1.00
    max_price = 10.00
    price_increment = 10.00
    total_price_ranges = (4000.00 - min_price) / price_increment

    q = queue.Queue()
    num_worker_threads = 5
    threads = []

    def update_progress(current_price):
        progress = int((current_price / 4000.00) * 100)
        cache.set("fetch_items_progress", {"progress": progress, "completed": False})

    # Start worker threads
    for _ in range(num_worker_threads):
        t = threading.Thread(target=process_queue, args=(q,))
        t.start()
        threads.append(t)

    try:
        while min_price <= 4000.00:
            current_page = 1
            total_pages = 1

            while current_page <= total_pages:
                try:
                    items, total_pages = fetch_finding_api_data(
                        page_number=current_page,
                        min_price=min_price,
                        max_price=max_price,
                    )

                    for item in items:
                        item_id = item["item_id"]
                        try:
                            if not Product.objects.filter(item_id=item_id).exists():
                                browse_data = fetch_browse_api_data(item_id)
                                combined_data = {**item, **browse_data}
                                q.put(combined_data)
                                total_items += 1
                            else:
                                print(
                                    f"product skipped {item_id} in page {current_page}, price range ${min_price}-${max_price}"
                                )
                        except Exception as e:
                            print(f"Error processing item {item_id}: {e}")

                except Exception as e:
                    print(
                        f"process:fetch_all_items Error fetching data for page {current_page}, price range ${min_price}-${max_price}: {e}"
                    )

                current_page += 1

            update_progress(min_price)
            min_price = max_price + 0.01
            max_price += price_increment

        # Wait for the queue to be processed
        q.join()

        # Stop worker threads
        for _ in range(num_worker_threads):
            q.put(None)
        for t in threads:
            t.join()

        # Generate HTML pages after fetching all items
        try:
            generate_html_pages_async.delay()
        except Exception as e:
            print(f"Error generating HTML pages: {e}")

        cache.set("fetch_items_progress", {"progress": 100, "completed": True})
        return JsonResponse(
            {
                "status": "success",
                "message": f"Fetched and stored information for {total_items} items",
                "total_items": total_items,
            },
            status=200,
        )

    except Exception as e:
        print(f"Error in fetch_all_items: {e}")
        cache.set("fetch_items_progress", {"progress": 0, "completed": True})
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@require_GET
def get_fetch_items_progress(request):
    progress_info = cache.get(
        "fetch_items_progress", {"progress": 0, "completed": False}
    )
    return JsonResponse(progress_info)


# def fetch_product_data(item_id):
#     finding_data = fetch_finding_api_data(item_id)
#     browse_data = fetch_browse_api_data(item_id)
#     return {**finding_data, **browse_data}

def fetch_finding_api_data(page_number=1, min_price=1.00, max_price=10.00):
    #print(f"\nFetching Finding API data:")
    #print(f"Page number: {page_number}")
    #print(f"Price range: ${min_price:.2f} - ${max_price:.2f}")
    
    headers = {
        "X-EBAY-SOA-SECURITY-APPNAME": "PJsAutoL-keyset-PRD-d2986db2a-865af761",
        "X-EBAY-SOA-OPERATION-NAME": "findItemsIneBayStores",
        "Content-Type": "text/xml",
    }
    #print("Headers prepared")

    xml_payload = f"""<?xml version="1.0" encoding="UTF-8"?>
    <findItemsIneBayStoresRequest xmlns="http://www.ebay.com/marketplace/search/v1/services">
        <storeName>PJ's Auto Literature</storeName>
        <outputSelector>StoreInfo</outputSelector>
        <itemFilter>
            <name>MinPrice</name>
            <value>{min_price:.2f}</value>
            <paramName>Currency</paramName>
            <paramValue>USD</paramValue>
        </itemFilter>
        <itemFilter>
            <name>MaxPrice</name>
            <value>{max_price:.2f}</value>
            <paramName>Currency</paramName>
            <paramValue>USD</paramValue>
        </itemFilter>
        <paginationInput>
            <pageNumber>{page_number}</pageNumber>
            <entriesPerPage>100</entriesPerPage>
        </paginationInput>
    </findItemsIneBayStoresRequest>"""
    #print("XML payload prepared")

    attempt = 0
    while attempt < MAX_RETRIES:
        try:
            #print(f"\nAttempt {attempt + 1} of {MAX_RETRIES}")
            #print("Sending API request...")
            response = requests.post(FINDING_API_URL, headers=headers, data=xml_payload)
            #print(f"Response status code: {response.status_code}")
            response.raise_for_status()

            #print("Parsing XML response...")
            root = ET.fromstring(response.content)
            namespace = {"ns": "http://www.ebay.com/marketplace/search/v1/services"}

            # Check for eBay API errors
            ack = root.find(".//ns:ack", namespace)
            #print(f"API acknowledgment: {ack.text if ack is not None else 'None'}")
            
            if ack is not None and ack.text != "Success":
                error_message = root.find(".//ns:errorMessage/ns:error/ns:message", namespace)
                error_code = root.find(".//ns:errorMessage/ns:error/ns:errorId", namespace)
                if error_message is not None and error_code is not None:
                    raise Exception(f"eBay API Error {error_code.text}: {error_message.text}")
                else:
                    raise Exception(f"Unknown eBay API Error: {ack.text}")

            items = root.findall(".//ns:item", namespace)
            #print(f"Found {len(items)} items in response")
            
            total_pages = int(root.find(".//ns:totalPages", namespace).text)
            #print(f"Total pages: {total_pages}")

            results = []
            for index, item in enumerate(items, 1):
                #print(f"\nProcessing item {index} of {len(items)}")
                try:
                    # Helper function to safely get text from XML element
                    def get_safe_text(element, path, default=None):
                        elem = element.find(path, namespace)
                        return elem.text if elem is not None else default

                    # Helper function to safely convert to boolean
                    def get_safe_bool(element, path):
                        return get_safe_text(element, path) == "true"

                    # Helper function to safely convert to integer
                    def get_safe_int(element, path):
                        text = get_safe_text(element, path)
                        return int(text) if text is not None else None

                    # Process price with rounding
                    price_elem = item.find("ns:sellingStatus/ns:currentPrice", namespace)
                    if price_elem is not None:
                        original_price = float(price_elem.text)
                        discounted_price = original_price * 0.97
                        # Using math.ceil to round up to 2 decimal places
                        rounded_price = math.ceil(discounted_price * 100) / 100
                        #print(f"Price processing:")
                        #print(f"  Original price: ${original_price:.2f}")
                        #print(f"  After discount: ${discounted_price:.2f}")
                        #print(f"  Final ceiling-rounded price: ${rounded_price:.2f}")
                    else:
                        rounded_price = 0.00

                    data = {
                        "item_id": get_safe_text(item, "ns:itemId"),
                        "title": get_safe_text(item, "ns:title"),
                        "global_id": get_safe_text(item, "ns:globalId"),
                        "category_id": get_safe_text(item, "ns:primaryCategory/ns:categoryId"),
                        "category_name": get_safe_text(item, "ns:primaryCategory/ns:categoryName"),
                        "gallery_url": get_safe_text(item, "ns:galleryURL"),
                        "view_item_url": get_safe_text(item, "ns:viewItemURL"),
                        "auto_pay": get_safe_bool(item, "ns:autoPay"),
                        "postal_code": get_safe_text(item, "ns:postalCode"),
                        "location": get_safe_text(item, "ns:location"),
                        "country": get_safe_text(item, "ns:country"),
                        "shipping_type": get_safe_text(item, "ns:shippingInfo/ns:shippingType"),
                        "ship_to_locations": get_safe_text(item, "ns:shippingInfo/ns:shipToLocations"),
                        "price": rounded_price,  # Using the rounded price
                        "expedited_shipping": get_safe_bool(item, "ns:shippingInfo/ns:expeditedShipping"),
                        "one_day_shipping_available": get_safe_bool(item, "ns:shippingInfo/ns:oneDayShippingAvailable"),
                        "handling_time": get_safe_int(item, "ns:shippingInfo/ns:handlingTime"),
                        "selling_state": get_safe_text(item, "ns:sellingStatus/ns:sellingState"),
                        "time_left": get_safe_text(item, "ns:sellingStatus/ns:timeLeft"),
                        "best_offer_enabled": get_safe_bool(item, "ns:listingInfo/ns:bestOfferEnabled"),
                        "buy_it_now_available": get_safe_bool(item, "ns:listingInfo/ns:buyItNowAvailable"),
                        "start_time": get_safe_text(item, "ns:listingInfo/ns:startTime"),
                        "end_time": get_safe_text(item, "ns:listingInfo/ns:endTime"),
                        "listing_type": get_safe_text(item, "ns:listingInfo/ns:listingType"),
                        "gift": get_safe_bool(item, "ns:listingInfo/ns:gift"),
                        "watch_count": get_safe_int(item, "ns:listingInfo/ns:watchCount"),
                        "returns_accepted": get_safe_bool(item, "ns:returnsAccepted"),
                        "is_multi_variation_listing": get_safe_bool(item, "ns:isMultiVariationListing"),
                        "top_rated_listing": get_safe_bool(item, "ns:topRatedListing")
                    }

                    # Remove None values to avoid serialization issues
                    data = {k: v for k, v in data.items() if v is not None}
                    #print(f"Successfully processed item {data.get('item_id', 'unknown')}")
                    results.append(data)

                except Exception as item_error:
                    #print(f"Error processing individual item {index}:")
                    #print(f"Error details: {str(item_error)}")
                    continue

            #print(f"\nSuccessfully processed {len(results)} items")
            return results, total_pages

        except requests.exceptions.RequestException as e:
            #print(f"HTTP Error (attempt {attempt + 1}): {e}")
            #print(f"Response status code: {response.status_code}")
            #print(f"Response headers: {response.headers}")
            #print(f"Response content: {response.content.decode('utf-8')}")
            pass
        except ET.ParseError as e:
            #print(f"XML Parsing Error (attempt {attempt + 1}): {e}")
            #print(f"Response content: {response.content.decode('utf-8')}")
            pass
        except Exception as e:
            #print(f"General Error (attempt {attempt + 1}): {e}")
            import traceback

            #print(f"Traceback: {traceback.format_exc()}")
            pass

        attempt += 1
        retry_delay = RETRY_DELAY * (2**attempt)
        #print(f"Retrying in {retry_delay} seconds...")
        time.sleep(retry_delay)  # Exponential backoff

    #print(f"Failed to fetch FindingService API data after {MAX_RETRIES} attempts")
    return [], 0


def fetch_new_access_token():
    url = "https://api.ebay.com/identity/v1/oauth2/token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Authorization": "Basic UEpzQXV0b0wta2V5c2V0LVBSRC1kMjk4NmRiMmEtODY1YWY3NjE6UFJELTI5ODZkYjJhMjg4NS01ZTE0LTQyNWMtOThlMS04Y2Fl",  # Replace with your actual credentials
    }
    data = {
        "grant_type": "refresh_token",
        "refresh_token": "v^1.1#i^1#f^0#p^3#I^3#r^1#t^Ul4xMF84OjNDRDY3NkNBOURFNzk4QzRFRDA1MTgzQTY5QTEzQTA3XzBfMSNFXjI2MA==",  # Replace with your actual refresh token
    }

    try:
        response = requests.post(url, headers=headers, data=data)
        response.raise_for_status()
        token_data = response.json()
        access_token = token_data.get("access_token")

        if access_token:
            # Update the EBAY_AUTH_TOKEN in settings
            settings.EBAY_AUTH_TOKEN = access_token
            return access_token
        else:
            raise ValueError("Failed to obtain access token.")
    except (RequestException, HTTPError) as e:
        print(f"Error fetching new access token: {e}")
        return None


# Global variables for rate limiting
RATE_LIMIT_PER_SECOND = 5
RATE_LIMIT_PERIOD = 1.0
token_bucket = RATE_LIMIT_PER_SECOND
last_request_time = time.time()
bucket_lock = Lock()


logger = logging.getLogger(__name__)


def fetch_browse_api_data(item_id):
    global token_bucket, last_request_time

    url = BROWSE_API_URL.format(item_id=item_id)

    headers = {
        "Authorization": f"Bearer {settings.EBAY_AUTH_TOKEN}",
        "X-EBAY-C-MARKETPLACE-ID": "EBAY_US",
        "X-EBAY-C-ENDUSERCTX": "sessionId=12345",
    }

    attempt = 0
    while attempt < MAX_RETRIES:
        try:
            with bucket_lock:
                current_time = time.time()
                time_passed = current_time - last_request_time
                token_bucket += time_passed * (
                    RATE_LIMIT_PER_SECOND / RATE_LIMIT_PERIOD
                )
                if token_bucket > RATE_LIMIT_PER_SECOND:
                    token_bucket = RATE_LIMIT_PER_SECOND

                if token_bucket < 1:
                    sleep_time = (1 - token_bucket) * (
                        RATE_LIMIT_PERIOD / RATE_LIMIT_PER_SECOND
                    )
                    time.sleep(sleep_time)
                    token_bucket = 1

                token_bucket -= 1
                last_request_time = time.time()

            response = requests.get(url, headers=headers)
            response.raise_for_status()
            data = response.json()

            if not data:
                logger.error(f"Empty response for item_id: {item_id}")
                return {}

            item_data = data.get("item", {})
            additional_images = data.get("additionalImages", [])
            additional_image_urls = [img["imageUrl"] for img in additional_images]
            
            result = {
                "price":data["price"]["value"],
                "description": data.get("description", ""),
                "currency": (
                    data["price"]["currency"]
                    if "price" in data and "currency" in data["price"]
                    else None
                ),
                "category_path": data.get("categoryPath", ""),
                "category_id_path": data.get("categoryIdPath", ""),
                "item_creation_date": data.get("itemCreationDate"),
                "estimated_availability_status": data.get(
                    "estimatedAvailabilities", [{}]
                )[0].get("estimatedAvailabilityStatus", ""),
                "estimated_available_quantity": data.get(
                    "estimatedAvailabilities", [{}]
                )[0].get("estimatedAvailableQuantity"),
                "estimated_sold_quantity": data.get("estimatedAvailabilities", [{}])[
                    0
                ].get("estimatedSoldQuantity"),
                "enabled_for_guest_checkout": data.get(
                    "enabledForGuestCheckout", False
                ),
                "eligible_for_inline_checkout": data.get(
                    "eligibleForInlineCheckout", False
                ),
                "lot_size": data.get("lotSize", 0),
                "legacy_item_id": data.get("legacyItemId", ""),
                "priority_listing": data.get("priorityListing", False),
                "adult_only": data.get("adultOnly", False),
                "listing_marketplace_id": data.get("listingMarketplaceId", ""),
                "shipping_service_code": data.get("shippingOptions", [{}])[0].get(
                    "shippingServiceCode", ""
                ),
                "shipping_carrier_code": data.get("shippingOptions", [{}])[0].get(
                    "shippingCarrierCode", ""
                ),
                "min_estimated_delivery_date": data.get("shippingOptions", [{}])[0].get(
                    "minEstimatedDeliveryDate"
                ),
                "max_estimated_delivery_date": data.get("shippingOptions", [{}])[0].get(
                    "maxEstimatedDeliveryDate"
                ),
                "shipping_cost": float(
                    data.get("shippingOptions", [{}])[0]
                    .get("shippingCost", {})
                    .get("value", 0)
                ),
                "shipping_cost_type": data.get("shippingOptions", [{}])[0].get(
                    "shippingCostType", ""
                ),
                "primary_image_url": (
                    data["image"]["imageUrl"]
                    if "image" in data and "imageUrl" in data["image"]
                    else None
                ),
                "additional_image_urls": additional_image_urls,
            }
            if "itemEndDate" in data:
                result["end_time"] = data["itemEndDate"]
            # Log any fields that are None
            none_fields = [field for field, value in result.items() if value is None]
            if none_fields:
                logger.warning(
                    f"Fields with None values for item_id {item_id}: {', '.join(none_fields)}"
                )
            print(f"[DEBUG] Successfully fetched data for item_id: {item_id}")
            return result

        except HTTPError as e:
            print(f"[ERROR] HTTP Error for item_id {item_id}: {str(e)}")
            if response.status_code == 401:
                print("[DEBUG] Attempting to fetch new access token")
                new_token = fetch_new_access_token()
                if new_token:
                    headers["Authorization"] = f"Bearer {new_token}"
                    continue
            attempt += 1
            print(f"[DEBUG] Retrying after {RETRY_DELAY * (2**attempt)} seconds")
            time.sleep(RETRY_DELAY * (2**attempt))

        except RequestException as e:
            print(f"[ERROR] Request Exception for item_id {item_id}: {str(e)}")
            attempt += 1
            time.sleep(RETRY_DELAY * (2**attempt))

    return {}


from django.utils.text import slugify


def save_product_data(product_data):
    try:
        sanitized_name = slugify(product_data["title"])
        base_name = sanitized_name
        counter = 1
        while Product.objects.filter(html_link=sanitized_name).exists():
            sanitized_name = f"{base_name}-{counter}"
            counter += 1

        description = product_data.get("description") or product_data.get(
            "short_description", ""
        )

        # print(f"description {description}")
        cleaned_description = clean_description(description)
    
        defaults = {
            "title": product_data["title"],
            "global_id": product_data.get("global_id"),
            "category_id": product_data.get("category_id"),
            "category_name": product_data.get("category_name"),
            "gallery_url": product_data.get("gallery_url"),
            "view_item_url": product_data.get("view_item_url"),
            "auto_pay": product_data.get("auto_pay", False),
            "postal_code": product_data.get("postal_code"),
            "location": product_data.get("location"),
            "country": product_data.get("country"),
            "selling_state": product_data.get("selling_state"),
            "time_left": product_data.get("time_left"),
            "best_offer_enabled": product_data.get("best_offer_enabled", False),
            "buy_it_now_available": product_data.get("buy_it_now_available", False),
            "start_time": product_data.get("start_time"),
            "end_time": product_data.get("end_time"),
            "listing_type": product_data.get("listing_type"),
            "gift": product_data.get("gift", False),
            "watch_count": product_data.get("watch_count"),
            "returns_accepted": product_data.get("returns_accepted", False),
            "is_multi_variation_listing": product_data.get(
                "is_multi_variation_listing", False
            ),
            "top_rated_listing": product_data.get("top_rated_listing", False),
            "short_description": cleaned_description,
            "price": product_data.get("price"),
            "currency": product_data.get("currency"),
            "category_path": product_data.get("category_path", ""),
            "category_id_path": product_data.get("category_id_path", ""),
            "item_creation_date": product_data.get("item_creation_date"),
            "estimated_availability_status": product_data.get(
                "estimated_availability_status", ""
            ),
            "estimated_available_quantity": product_data.get(
                "estimated_available_quantity"
            ),
            "estimated_sold_quantity": product_data.get("estimated_sold_quantity"),
            "enabled_for_guest_checkout": product_data.get(
                "enabled_for_guest_checkout", False
            ),
            "eligible_for_inline_checkout": product_data.get(
                "eligible_for_inline_checkout", False
            ),
            "lot_size": product_data.get("lot_size", 0),
            "legacy_item_id": product_data.get("legacy_item_id", ""),
            "priority_listing": product_data.get("priority_listing", False),
            "adult_only": product_data.get("adult_only", False),
            "listing_marketplace_id": product_data.get("listing_marketplace_id", ""),
            "seller_username": product_data.get("seller_username"),
            "feedback_score": product_data.get("feedback_score"),
            "positive_feedback_percent": product_data.get("positive_feedback_percent"),
            "feedback_rating_star": product_data.get("feedback_rating_star"),
            "top_rated_seller": product_data.get("top_rated_seller", False),
            "shipping_type": product_data.get("shipping_type"),
            "ship_to_locations": product_data.get("ship_to_locations"),
            "expedited_shipping": product_data.get("expedited_shipping", False),
            "one_day_shipping_available": product_data.get(
                "one_day_shipping_available", False
            ),
            "handling_time": product_data.get("handling_time"),
            "shipping_service_code": product_data.get("shipping_service_code"),
            "shipping_carrier_code": product_data.get("shipping_carrier_code"),
            "min_estimated_delivery_date": product_data.get(
                "min_estimated_delivery_date"
            ),
            "max_estimated_delivery_date": product_data.get(
                "max_estimated_delivery_date"
            ),
            "shipping_cost": product_data.get("shipping_cost"),
            "shipping_cost_type": product_data.get("shipping_cost_type"),
            "primary_image_url": product_data.get("primary_image_url"),
            "additional_image_urls": product_data.get("additional_image_urls"),
            "html_link": sanitized_name,
        }

        product, created = Product.objects.update_or_create(
            item_id=product_data["item_id"], defaults=defaults
        )

        if created:
            logger.info(
                f"Created new product: {product_data['title']} - {product_data['item_id']}"
            )
        else:
            logger.info(
                f"Updated existing product: {product_data['title']} - {product_data['item_id']}"
            )

    except IntegrityError as e:
        logger.error(
            f"IntegrityError while saving product {product_data['item_id']}: {e}"
        )
    except Exception as e:
        logger.error(f"Error saving product data: {e}")
        logger.error(f"Product data: {product_data['item_id']}")


import ast
import logging
import os

from celery import shared_task
from django.conf import settings
from django.core.cache import cache
from django.http import JsonResponse
from django.template import loader
from django.views.decorators.http import require_GET

from .models import \
    Product  # Ensure this import is correct for your project structure

logger = logging.getLogger(__name__)

@shared_task
def generate_html_pages_async():
    # Check if the generation is already in progress

    # Set the in-progress flag

    output_dir = os.path.join(settings.BASE_DIR, "products", "viewproduct")
    os.makedirs(output_dir, exist_ok=True)

    products = Product.objects.all()
    total_products = products.count()
    completed = 0
    errors = []

    try:
        template = loader.get_template("pages/template.html")
    except Exception as e:
        logger.error(f"Error loading template: {e}")
        cache.set("html_generation_in_progress", False)
        return
    start_time = time.time()
    for product in products:
        if product.additional_image_urls:
            try:
                additional_images = ast.literal_eval(product.additional_image_urls)
                additional_images = [
                    str(url).strip() for url in additional_images if url
                ]
                
            except Exception as e:
                logger.error(
                    f"Error parsing additional image URLs for product {product.item_id}: {e}"
                )
                additional_images = []
        else:
            additional_images = []

        context = {"product": product, "additional_images": additional_images}

        try:
            rendered_html = template.render(context)
        except Exception as e:
            errors.append(f"Error rendering HTML for product {product.item_id}: {e}")
            continue

        filename = f"{product.html_link}.html"
        filepath = os.path.join(output_dir, filename)

        try:
            with open(filepath, "w", encoding="utf-8") as file:
                file.write(rendered_html)
        except Exception as e:
            errors.append(f"Error writing file for product {product.item_id}: {e}")

        # Update progress
        completed += 1
        elapsed_time = time.time() - start_time
        progress = int((completed / total_products) * 100) if total_products > 0 else 100
        items_per_second = completed / elapsed_time if elapsed_time > 0 else 0

    # Mark as completed
    

    if errors:
        logger.error(f"Errors encountered while generating HTML pages: {errors}")
    else:
        logger.info("HTML pages generated successfully.")

import time

from django.http import JsonResponse
from django.views.decorators.http import require_GET

from .models import Product


@require_GET
def run_html_generation(request):
    try:
        generate_html_pages_async.delay()
        return JsonResponse({"status": "success", "message": "HTML generation task has been scheduled"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

import time

from django.db.models import Count

from .models import Product


def get_html_generation_progress(request):
    start_time = time.time()
    product_count = Product.objects.count()
    
    # Fetch the first 100 products to estimate the processing rate
    initial_products = Product.objects.all()[:100]
    
    # Count the number of products processed in the first 5 seconds
    for i, _ in enumerate(initial_products):
        if time.time() - start_time >= 5:
            break
    
    # Calculate the items processed per second
    items_per_second = i    
    # Get the total number of products
    total_products = Product.objects.count()
    
    return JsonResponse({
        "items_per_second": items_per_second,
        "total": total_products
    })



import json

from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render
from django.utils import timezone

from .models import CalendarEvent, CartItem


def landing_page(request):
    cart_count = 0

    # Check if there's a cart in the session
    cart_id = request.session.get("cart_id")
    if cart_id:
        try:
            # Get all cart items related to the cart
            cart_items = CartItem.objects.filter(cart_id=cart_id)
            # Sum up the quantity of all items in the cart to get the cart count
            cart_count = sum(item.quantity for item in cart_items)
        except CartItem.DoesNotExist:
            cart_count = 0

    # Fetch calendar events
    events = CalendarEvent.objects.all().order_by("start_date")

    # Convert events to a dictionary format with both start and end dates
    events_dict = {
        str(event.id): {
            'title': event.title,
            'description': event.description,
            'location': event.location,
            'start_date': event.start_date.isoformat(),
            'end_date': event.end_date.isoformat()
        }
        for event in events
    }

    events_json = json.dumps(events_dict, default=str)

    return render(
        request,
        "pages/landing.html",
        {
            "cart_count": cart_count,
            "events_json": events_json,
        },
    )

def terms_view(request):
    cart_count = 0

    # Check if there's a cart in the session
    cart_id = request.session.get("cart_id")
    if cart_id:
        try:
            # Get all cart items related to the cart
            cart_items = CartItem.objects.filter(cart_id=cart_id)
            # Sum up the quantity of all items in the cart to get the cart count
            cart_count = sum(item.quantity for item in cart_items)
        except CartItem.DoesNotExist:
            # If the cart items don't exist, the cart is empty or invalid
            cart_count = 0
    return render(
        request,
        "pages/terms.html",
        {
            "cart_count": cart_count,
        },
    )
def contact_view(request):
    cart_count = 0

    # Check if there's a cart in the session
    cart_id = request.session.get("cart_id")
    if cart_id:
        try:
            # Get all cart items related to the cart
            cart_items = CartItem.objects.filter(cart_id=cart_id)
            # Sum up the quantity of all items in the cart to get the cart count
            cart_count = sum(item.quantity for item in cart_items)
        except CartItem.DoesNotExist:
            # If the cart items don't exist, the cart is empty or invalid
            cart_count = 0
    return render(
        request,
        "pages/contact.html",
        {
            "cart_count": cart_count,
        },
    )


def admin_page(request):
    return render(request, "pages/admin.html")


def admin_page2(request):
    return render(request, "pages/admin-2.html")


def admin_page3(request):
    return render(request, "pages/admin-3.html")

from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string

from .models import Order, OrderItem, ShippingAddress


def order_list(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'pages/orders.html', {'orders': orders})

def order_details(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order_items = order.items.all()
    shipping_address = order.shipping_address

    context = {
        'order': order,
        'order_items': order_items,
        'shipping_address': shipping_address
    }

    # Render details as an HTML fragment for the modal
    html = render_to_string('pages/order_details.html', context)
    return HttpResponse(html)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .models import Order


def check_order_status(request, order_id):
    try:
        order = Order.objects.get(order_number=order_id)
        return JsonResponse({'status': 'success', 'delivery_status': order.get_delivery_status_display()})
    except Order.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Order not found'})

@csrf_exempt
def update_delivery_status(request, order_id):
    if request.method == 'POST':
        try:
            order = Order.objects.get(id=order_id)
            data = json.loads(request.body)
            order.delivery_status = data['delivery_status']
            order.save()
            return JsonResponse({'status': 'success'})
        except Order.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Order not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
# views.py


from django.db.models import Q


def product_list(request):
    query = request.GET.get("query", "")
    products = Product.objects.all()

    cart_count = 0
    cart_id = request.session.get("cart_id")

    if cart_id:
        try:
            # Get all cart items related to the cart
            cart_items = CartItem.objects.filter(cart_id=cart_id)
            # Sum up the quantity of all items in the cart to get the cart count
            cart_count = sum(item.quantity for item in cart_items)
        except CartItem.DoesNotExist:
            # If the cart items don't exist, the cart is empty or invalid
            cart_count = 0

    if query:
        # Split the query into words and create a Q object to search across title and short_description
        query_words = query.split()
        q_objects = Q()
        
        for word in query_words:
            q_objects &= (Q(title__icontains=word) | Q(short_description__icontains=word))
        
        # Apply the filter to get products that match all words in the query
        products = products.filter(q_objects)

    paginator = Paginator(products, 10)  # Show 10 products per page.
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "query": query,
        "cart_count": cart_count,
    }
    return render(request, "pages/product_list.html", context)



def cron(request):
    return HttpResponse("Happy")


def clean_description(description):
    # Check if the description contains HTML
    if "<" in description and ">" in description:
        # Parse the HTML and extract the text content
        soup = BeautifulSoup(description, "html.parser")
        text_content = soup.get_text()
    else:
        # If it's not HTML, use the description as is
        text_content = description

    # Remove unwanted characters (like Â from encoding issues)
    cleaned_text = text_content.replace("Â", "")

    # Remove everything after the first occurrence of the word "condition"
    cleaned_text = (
        re.split(r"(condition)\b", cleaned_text, flags=re.IGNORECASE)[0] + "condition."
    )

    # Remove extra whitespace
    cleaned_text = re.sub(r"\s+", " ", cleaned_text).strip()

    return cleaned_text


import threading


def clean_short_description(description):
    if not description:
        return description
    cleaned_text = (
        re.split(r"(condition)\b.*", description, flags=re.IGNORECASE)[0] + " condition"
    )
    return cleaned_text.strip()


def process_product(q, updated_count, updated_count_lock):
    while True:
        product = q.get()
        if product is None:
            break

        original_description = product.short_description
        cleaned_description = clean_short_description(original_description)

        if original_description != cleaned_description:
            product.short_description = cleaned_description
            product.save()

            with updated_count_lock:
                updated_count[0] += 1

        q.task_done()


@require_GET
def fetch_and_print_descriptions(request):
    try:
        products = Product.objects.exclude(short_description__isnull=True)
        updated_count = [0]  # Use a list to allow modification within threads
        updated_count_lock = threading.Lock()

        q = queue.Queue()

        # Start worker threads
        threads = []
        for _ in range(30):
            t = threading.Thread(
                target=process_product, args=(q, updated_count, updated_count_lock)
            )
            t.start()
            threads.append(t)

        # Queue the products for processing
        for product in products:
            q.put(product)

        # Wait for the queue to be processed
        q.join()

        # Stop worker threads
        for _ in range(30):
            q.put(None)
        for t in threads:
            t.join()

        return JsonResponse(
            {
                "status": "success",
                "message": f"Cleaned and updated {updated_count[0]} descriptions.",
            },
            status=200,
        )

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


from concurrent.futures import ThreadPoolExecutor, as_completed

MAX_WORKERS = 30


def process_price_range(
    min_price, max_price, existing_item_ids, updated_item_ids_queue, fields_to_check
):
    print(f"Processing price range: ${min_price:.2f} - ${max_price:.2f}")
    current_page = 1
    total_pages = 1
    local_updated_item_ids = set()

    # Field mapping for mismatched names
    field_mapping = {
        "description": "short_description",
        # Add any other mismatched field names here
    }

    while current_page <= total_pages:
        try:
            items, total_pages = fetch_finding_api_data(
                page_number=current_page, min_price=min_price, max_price=max_price
            )
            print(
                f"Fetched page {current_page} of {total_pages} for price range ${min_price:.2f} - ${max_price:.2f}"
            )

            for item in items:
                item_id = item["item_id"]
                local_updated_item_ids.add(item_id)

                # Apply field mapping
                mapped_item = {
                    field_mapping.get(k, k): v
                    for k, v in item.items()
                    if k in fields_to_check
                }

                try:
                    product = Product.objects.get(item_id=item_id)
                    before_dict = {
                        key: getattr(product, key)
                        for key in fields_to_check
                        if hasattr(product, key)
                    }
                    after_dict = {
                        key: value
                        for key, value in mapped_item.items()
                        if key in fields_to_check and hasattr(product, key)
                    }

                    changes = {
                        key: value
                        for key, value in after_dict.items()
                        if before_dict.get(key) != value
                    }

                    if changes:
                        for key, value in changes.items():
                            setattr(product, key, value)
                        product.save()
                        print(f"Updated product: {item_id} - {product.title}")
                        print(f"Changes: {changes}")

                        change_log = ProductChangeLog.objects.create(
                            item_id=item_id,
                            product_name=product.title,
                            operation="updated",
                        )
                        change_log.set_changes(before_dict, after_dict)
                        change_log.save()
                    else:
                        print(
                            f"No changes for checked fields in product: {item_id} - {product.title}"
                        )
                except Product.DoesNotExist:
                    browse_data = fetch_browse_api_data(item_id)
                    mapped_browse_data = {
                        field_mapping.get(k, k): v
                        for k, v in browse_data.items()
                        if k in fields_to_check
                    }
                    combined_data = {**mapped_item, **mapped_browse_data}
                    new_product = Product.objects.create(**combined_data)
                    print(f"Created new product: {item_id} - {new_product.title}")
                    ProductChangeLog.objects.create(
                        item_id=item_id,
                        product_name=new_product.title,
                        operation="created",
                    )

            current_page += 1

        except Exception as e:
            print(
                f"process: process_price_range Error fetching data for page {current_page}, price range ${min_price:.2f} - ${max_price:.2f}: {e}"
            )

    updated_item_ids_queue.put(local_updated_item_ids)

import json
import time
from decimal import Decimal

from celery import shared_task
from celery.exceptions import Ignore
from celery.result import AsyncResult
from django.core.serializers.json import DjangoJSONEncoder
from django.http import JsonResponse, StreamingHttpResponse
from django.utils import timezone


class DecimalEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

@shared_task(bind=True)
def run_daily_update_async(self):
    #print("=" * 50)
    #print("Starting daily update...")
    #print("Initializing task state...")
    
    def is_aborted():
        abort_status = (
            self.request.called_directly
            or self.request.id is None
            or self.AsyncResult(self.request.id).state == "REVOKED"
        )
        #print(f"Checking abort status: {abort_status}")
        return abort_status

    try:
        #print("Creating/getting fetch status...")
        fetch_status, created = FetchStatus.objects.get_or_create(fetch_type="daily")
        #print(f"Fetch status created: {created}")

        products_processed = 0
        #print("Getting existing item IDs...")
        existing_item_ids = set(Product.objects.values_list("item_id", flat=True))
        #print(f"Found {len(existing_item_ids)} existing items")
        updated_item_ids = set()

        #print("Setting up price ranges...")
        price_ranges = []
        min_price = Decimal('1.00')
        max_price = Decimal('10.00')
        price_increment = Decimal('10.00')

        while min_price <= Decimal('4000.00'):
            price_ranges.append((float(min_price), float(max_price)))  # Convert to float
            #print(f"Added price range: ${min_price:.2f} - ${max_price:.2f}")
            min_price = max_price + Decimal('0.01')
            max_price += price_increment

        #print(f"Created {len(price_ranges)} price ranges")

        fields_to_check = [
            "item_id", "title", "global_id", "category_id", "category_name",
            "gallery_url", "view_item_url", "auto_pay", "postal_code",
            "location", "country", "shipping_type", "ship_to_locations", "price"
        ]
        #print(f"Fields to check: {fields_to_check}")

        field_mapping = {"description": "short_description"}
        
        try:
            for min_price, max_price in price_ranges:
                #print(f"\nProcessing price range: ${min_price:.2f} - ${max_price:.2f}")
                if is_aborted():
                    #print("Task aborted during price range processing")
                    raise Ignore()

                current_page = 1
                total_pages = 1

                while current_page <= total_pages:
                    #print(f"\nProcessing page {current_page} of {total_pages}")
                    if is_aborted():
                        #print("Task aborted during page processing")
                        raise Ignore()

                    try:
                        #print(f"Fetching data for page {current_page}...")
                        items, total_pages = fetch_finding_api_data(
                            page_number=current_page,
                            min_price=min_price,
                            max_price=max_price,
                        )
                        #print(f"Fetched {len(items)} items from API")

                        for item in items:
                            item_id = item["item_id"]
                            #print(f"\nProcessing item: {item_id}")
                            updated_item_ids.add(item_id)

                            # Convert Decimal to float in the item data
                            for key, value in item.items():
                                if isinstance(value, Decimal):
                                    item[key] = float(value)
                                    #print(f"Converted Decimal to float for {key}: {value}")

                            mapped_item = {
                                field_mapping.get(k, k): v for k, v in item.items()
                            }

                            try:
                                #print(f"Checking if product exists: {item_id}")
                                product = Product.objects.get(item_id=item_id)
                                #print("Product found, checking for changes...")
                                
                                before_dict = {
                                    key: getattr(product, key)
                                    for key in fields_to_check
                                    if hasattr(product, key)
                                }
                                
                                # Convert Decimal values to float in before_dict
                                before_dict = {
                                    k: float(v) if isinstance(v, Decimal) else v
                                    for k, v in before_dict.items()
                                }
                                
                                after_dict = {
                                    key: value
                                    for key, value in mapped_item.items()
                                    if key in fields_to_check and hasattr(product, key)
                                }

                                changes = {
                                    key: value
                                    for key, value in after_dict.items()
                                    if before_dict.get(key) != value
                                }

                                if changes:
                                    #print(f"Changes detected for {item_id}:")
                                    for key, value in changes.items():
                                        #print(f"  {key}: {before_dict.get(key)} -> {value}")
                                        setattr(product, key, value)
                                    product.save()

                                    change_log = ProductChangeLog.objects.create(
                                        item_id=item_id,
                                        product_name=product.title,
                                        operation="updated",
                                    )
                                    change_log.set_changes(before_dict, after_dict)
                                    change_log.save()
                                    #print("Change log created")
                                #else:
                                    #print("No changes detected")

                            except Product.DoesNotExist:
                                #print(f"Product {item_id} not found, creating new...")
                                browse_data = fetch_browse_api_data(item_id)
                                
                                # Convert Decimal values in browse_data
                                for key, value in browse_data.items():
                                    if isinstance(value, Decimal):
                                        browse_data[key] = float(value)
                                        #print(f"Converted Decimal to float in browse_data: {key}")
                                
                                mapped_browse_data = {
                                    field_mapping.get(k, k): v
                                    for k, v in browse_data.items()
                                }
                                combined_data = {**mapped_item, **mapped_browse_data}
                                save_product_data(combined_data)
                                #print(f"Created new product: {item_id}")

                                ProductChangeLog.objects.create(
                                    item_id=item_id,
                                    product_name=combined_data.get("title", "Unknown Title"),
                                    operation="created",
                                )
                                #print("Creation log created")

                            products_processed += 1
                            #print(f"Products processed: {products_processed}")
                            self.update_state(
                                state="PROGRESS",
                                meta={"products_processed": products_processed},
                            )

                        current_page += 1

                    except Exception as e:
                        #print(f"Error processing page {current_page}:")
                        #print(f"Error details: {str(e)}")
                        #print(f"Error type: {type(e)}")
                        import traceback

                        #print(f"Traceback: {traceback.format_exc()}")
                        pass

        except Exception as e:
            #print(f"Error during main processing loop:")
            #print(f"Error details: {str(e)}")
            return {"status": "error", "message": str(e)}

        #print("\nProcessing deletions...")
        items_to_delete = existing_item_ids - updated_item_ids
        #print(f"Found {len(items_to_delete)} items to check for deletion")

        for item_id in items_to_delete:
            #print(f"\nChecking item {item_id} for deletion...")
            if is_aborted():
                #print("Task aborted during deletion process")
                raise Ignore()

            listing_status = get_ebay_listing_status(item_id)
            #print(f"eBay listing status: {listing_status}")
            
            if listing_status != "Active":
                try:
                    product = Product.objects.get(item_id=item_id)
                    product_name = product.title
                    product.delete()
                    #print(f"Deleted inactive product: {item_id} - {product_name}")
                    
                    ProductChangeLog.objects.create(
                        item_id=item_id,
                        product_name=product_name,
                        operation="deleted"
                    )
                    #print("Deletion log created")
                except Product.DoesNotExist:
                    #print(f"Product {item_id} not found in database, skipping deletion")
                    pass

            products_processed += 1
            #print(f"Total products processed: {products_processed}")
            self.update_state(
                state="PROGRESS",
                meta={"products_processed": products_processed}
            )

        #print("\nUpdating fetch status...")
        fetch_status.last_run = timezone.now()
        fetch_status.save()
        
        #print("Triggering HTML page generation...")
        generate_html_pages_async.delay()

        #print("Daily update completed successfully.")
        return {"status": "completed", "products_processed": products_processed}
        
    except Exception as e:
        #print("Fatal error in daily update:")
        #print(f"Error type: {type(e)}")
        #print(f"Error message: {str(e)}")
        import traceback

        #print(f"Traceback: {traceback.format_exc()}")
        return {"status": "error", "message": str(e)}
    
import time

from celery.result import AsyncResult
# views.py
from django.http import JsonResponse
from django.utils import timezone

from .models import Product


def run_daily_update(request):
    try:
        task = run_daily_update_async.delay()
        return JsonResponse({
            "status": "success",
            "task_id": task.id,
            "message": "Daily update task has been scheduled"
        })
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

def update_progress(request, task_id):
    try:
        task = AsyncResult(task_id)
        
        # Get the start time of the task
        if not hasattr(task, 'start_time'):
            task.start_time = time.time()
        
        current_time = time.time()
        elapsed_time = current_time - task.start_time
        
        # Get current progress from task
        if isinstance(task.info, dict):
            products_processed = task.info.get('products_processed', 0)
        else:
            products_processed = 0
            
        # Calculate processing rate (items per second)
        processing_rate = products_processed / elapsed_time if elapsed_time > 0 else 0
        
        # Get total number of products
        total_products = Product.objects.count()
        
        return JsonResponse({
            'status': 'success',
            'processing_rate': round(processing_rate, 2),  # items per second
            'total_products': total_products,
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

from celery.result import AsyncResult


def cancel_update(request):
    task_id = request.GET.get("task_id")
    if task_id:
        task = AsyncResult(task_id)
        task.revoke(terminate=True)
        return JsonResponse(
            {"success": True, "message": "Update task has been canceled"}
        )
    else:
        return JsonResponse({"success": False, "error": "No task ID provided"})


# Add any other view functions here

# @require_GET
# def run_daily_update(request):
#     cache.set('generate_report_progress', {'progress': 0, 'completed': False})

#     try:
#         fetch_status, created = FetchStatus.objects.get_or_create(fetch_type='daily')
#         existing_item_ids = set(Product.objects.values_list('item_id', flat=True))
#         updated_item_ids = set()

#         price_ranges = []
#         min_price = 1.00
#         max_price = 10.00
#         price_increment = 10.00

#         while min_price <= 4000.00:
#             price_ranges.append((min_price, max_price))
#             min_price = max_price + 0.01
#             max_price += price_increment

#         fields_to_check = ['item_id', 'title', 'global_id', 'category_id', 'category_name', 'gallery_url',
#                            'view_item_url', 'auto_pay', 'postal_code', 'location', 'country', 'shipping_type',
#                            'ship_to_locations']
#         field_mapping = {
#             'description': 'short_description',
#         }

#         for index, (min_price, max_price) in enumerate(price_ranges):
#             current_page = 1
#             total_pages = 1

#             while current_page <= total_pages:
#                 try:
#                     items, total_pages = fetch_finding_api_data(page_number=current_page, min_price=min_price, max_price=max_price)

#                     for item in items:
#                         item_id = item['item_id']
#                         updated_item_ids.add(item_id)

#                         # Apply field mapping to the fetched item
#                         mapped_item = {field_mapping.get(k, k): v for k, v in item.items() if k in fields_to_check}

#                         try:
#                             product = Product.objects.get(item_id=item_id)
#                             before_dict = {key: getattr(product, key) for key in fields_to_check if hasattr(product, key)}
#                             after_dict = {key: value for key, value in mapped_item.items() if key in fields_to_check and hasattr(product, key)}

#                             changes = {key: value for key, value in after_dict.items() if before_dict.get(key) != value}

#                             if changes:
#                                 # Print the product ID and what has changed
#                                 print(f"Updating product {item_id}: {changes}", flush=True)

#                                 for key, value in changes.items():
#                                     setattr(product, key, value)
#                                 product.save()

#                                 change_log = ProductChangeLog.objects.create(
#                                     item_id=item_id,
#                                     product_name=product.title,
#                                     operation='updated'
#                                 )
#                                 change_log.set_changes(before_dict, after_dict)
#                                 change_log.save()
#                         except Product.DoesNotExist:
#                             browse_data = fetch_browse_api_data(item_id)
#                             # Apply field mapping to the browse data
#                             mapped_browse_data = {field_mapping.get(k, k): v for k, v in browse_data.items() if k in fields_to_check}
#                             combined_data = {**mapped_item, **mapped_browse_data}
#                             new_product = Product.objects.create(**combined_data)
#                             print(f"Created new product {item_id}", flush=True)  # Print new product creation
#                             ProductChangeLog.objects.create(
#                                 item_id=item_id,
#                                 product_name=new_product.title,
#                                 operation='created'
#                             )

#                     current_page += 1

#                 except Exception as e:
#                     print(f"Error fetching data for page {current_page}, price range ${min_price:.2f}-${max_price:.2f}: {e}", flush=True)

#             # Update progress
#             progress = int(((index + 1) / len(price_ranges)) * 100)
#             cache.set('generate_report_progress', {'progress': progress, 'completed': False})

#         deleted_item_ids = existing_item_ids - updated_item_ids
#         for item_id in deleted_item_ids:
#             product = Product.objects.get(item_id=item_id)
#             product_name = product.title
#             product.delete()
#             print(f"Deleted product {item_id}", flush=True)  # Print product deletion
#             ProductChangeLog.objects.create(
#                 item_id=item_id,
#                 product_name=product_name,
#                 operation='deleted'
#             )

#         fetch_status.last_run = timezone.now()
#         fetch_status.save()

#         cache.set('generate_report_progress', {'progress': 100, 'completed': True})
#         return JsonResponse({"status": "success", "message": "Daily update completed successfully"})

#     except Exception as e:
#         print(f"Error in run_daily_update: {e}", flush=True)
#         cache.set('generate_report_progress', {'progress': 0, 'completed': True})
#         return JsonResponse({"status": "error", "message": str(e)})


import datetime
import io
# Generate report only:
import json
import logging
import time
from datetime import datetime

from celery import shared_task
from django.http import StreamingHttpResponse
from django.utils import timezone

from .models import FetchStatus, Product, Report

logger = logging.getLogger(__name__)


class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%d %H:%M:%S")
        return super().default(obj)


@shared_task(bind=True)
def generate_report_async(self):
    logger.info("Starting the report generation task.")
    self.update_state(state="PROGRESS", meta={"progress": 0})

    fetch_status, created = FetchStatus.objects.get_or_create(fetch_type="report")

    total_products = Product.objects.count()
    processed_items = 0

    existing_item_ids = set(Product.objects.values_list("item_id", flat=True))
    updated_item_ids = set()

    price_ranges = []
    min_price = 1.00
    max_price = 10.00
    price_increment = 10.00

    while min_price <= 4000.00:
        price_ranges.append((min_price, max_price))
        min_price = max_price + 0.01
        max_price += price_increment

    fields_to_check = [
        "title", "global_id", "category_id", "category_name",
        "gallery_url", "view_item_url", "auto_pay", "postal_code",
        "location", "country", "shipping_type", "ship_to_locations", 
        "price", "short_description"
    ]

    field_mapping = {
        "description": "short_description",
    }

    reports = []

    try:
        for min_price, max_price in price_ranges:
            logger.info(f"Processing price range: ${min_price:.2f} - ${max_price:.2f}")
            current_page = 1
            range_total_pages = 1

            while current_page <= range_total_pages:
                try:
                    logger.info(
                        f"Fetching data for page {current_page}, price range ${min_price:.2f} - ${max_price:.2f}"
                    )
                    items, range_total_pages = fetch_finding_api_data(
                        page_number=current_page,
                        min_price=min_price,
                        max_price=max_price,
                    )

                    for item in items:
                        item_id = item["item_id"]
                        updated_item_ids.add(item_id)

                        mapped_item = {
                            field_mapping.get(k, k): v for k, v in item.items()
                        }

                        product = Product.objects.filter(item_id=item_id).first()
                        if product:
                            before_dict = {
                                key: getattr(product, key)
                                for key in fields_to_check
                                if hasattr(product, key)
                            }
                            after_dict = {
                                key: value
                                for key, value in mapped_item.items()
                                if key in fields_to_check and hasattr(product, key)
                            }

                            changes = {
                                key: value
                                for key, value in after_dict.items()
                                if before_dict.get(key) != value
                            }

                            if changes:
                                logger.info(
                                    f"Changes detected for product: {item_id} - {product.title}"
                                )
                                report = Report(
                                    item_id=item_id,
                                    product_name=product.title,
                                    operation="changes_detected",
                                )
                                report.set_changes(before_dict, after_dict)
                                reports.append(report)
                        else:
                            logger.info(
                                f"New product detected: {item_id} - {mapped_item.get('title')}"
                            )
                            report = Report(
                                item_id=item_id,
                                product_name=mapped_item.get("title"),
                                operation="new_product_detected",
                                changes=None,  # Set changes to None for new products
                            )
                            reports.append(report)

                        processed_items += 1
                        progress = int(processed_items / total_products * 100)
                        self.update_state(state="PROGRESS", meta={"progress": progress})

                    current_page += 1

                except Exception as e:
                    logger.error(
                        f"process: generate_report Error fetching data for page {current_page}, price range ${min_price:.2f} - ${max_price:.2f}: {e}"
                    )

    except Exception as e:
        logger.error(f"Error during report generation: {e}")

    # Check for potentially deleted items
    items_to_delete = existing_item_ids - updated_item_ids
    for item_id in items_to_delete:
        product = Product.objects.filter(item_id=item_id).first()
        if product:
            report = Report(
                item_id=item_id,
                product_name=product.title,
                operation="potential_deletion",
                changes=None,  # Set changes to None for deleted products
            )
            reports.append(report)
            logger.info(f"Potential deletion detected: {item_id} - {product.title}")

    # Save all reports to the database
    if reports:
        Report.objects.bulk_create(reports)

    fetch_status.last_run = timezone.now()
    fetch_status.save()

    logger.info("Report generation task completed successfully.")
    return {"status": "completed", "progress": 100}


def generate_report(request):
    task = generate_report_async.delay()
    return JsonResponse({"task_id": task.id})


from django.http import StreamingHttpResponse


def report_progress(request, task_id):
    response = StreamingHttpResponse(
        event_stream(task_id), content_type="text/event-stream"
    )
    response["Cache-Control"] = "no-cache"
    return response


def event_stream(task_id):
    task = generate_report_async.AsyncResult(task_id)
    previous_progress = 0

    while not task.ready():
        # Get the current state
        current_state = task.info  # task.info holds the meta data from update_state
        if current_state and isinstance(current_state, dict):
            progress = current_state.get("progress", 0)

            if progress != previous_progress:
                yield f"data: {json.dumps({'progress': progress})}\n\n"
                previous_progress = progress

        time.sleep(1)  # Adjust polling interval

    # Final completion message
    yield f"data: {json.dumps({'status': 'completed'})}\n\n"


import datetime
from datetime import datetime
from datetime import time as time_module  # Importing both datetime and time

# or you can use 'import datetime' and then refer to datetime.datetime below
from django.db.models import Max
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET

from .models import Report


@require_GET
def fetch_report_log(request):
    try:
        print("Starting fetch_report_log")

        date_str = request.GET.get("date")
        print(f"Received date_str: {date_str}")

        if date_str:
            try:
                selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                print(f"Parsed selected_date: {selected_date}")
            except ValueError as ve:
                print(f"Error parsing date: {str(ve)}")
                return JsonResponse({"error": "Invalid date format"}, status=400)
        else:
            # Fetch the latest report date if no date is provided
            latest_log_date = Report.objects.aggregate(Max("date"))["date__max"]
            print(f"Latest log date: {latest_log_date}")
            if latest_log_date is None:
                print("No report entries found")
                return JsonResponse({"error": "No report entries found"}, status=404)
            selected_date = latest_log_date.date()
            print(f"Using latest selected_date: {selected_date}")

        try:
            # Use time.min and time.max instead of datetime.min and datetime.max
            start_date = timezone.make_aware(
                datetime.combine(selected_date, time_module.min)
            )
            end_date = timezone.make_aware(
                datetime.combine(selected_date, time_module.max)
            )
            print(f"Start date: {start_date}, End date: {end_date}")
        except Exception as dt_err:
            print(f"Error creating start or end date: {str(dt_err)}")
            return JsonResponse({"error": "Error processing date"}, status=500)

        # Fetch all reports for the selected date
        try:
            reports = Report.objects.filter(date__range=(start_date, end_date)).order_by(
                "-date"
            )
            print(f"Number of reports found: {reports.count()}")
        except Exception as db_err:
            print(f"Error querying reports: {str(db_err)}")
            return JsonResponse({"error": "Error fetching reports"}, status=500)

        data = []
        for report in reports:
            try:
                report_data = {
                    "item_id": report.item_id,
                    "product_name": report.product_name,
                    "operation": report.get_operation_display(),
                    "date": report.date.isoformat(),
                    "changes": report.changes,  # Assuming 'changes' is already a dict
                }
                data.append(report_data)
            except Exception as e:
                print(f"Error processing report entry {report.id}: {str(e)}")

        print(f"Successfully processed {len(data)} reports")

        return JsonResponse(
            {
                "date": selected_date.isoformat(),
                "data": data,
            },
            safe=False,
        )

    except Exception as e:
        print(f"Unexpected error in fetch_report_log: {str(e)}", exc_info=True)
        return JsonResponse({"error": "An unexpected error occurred"}, status=500)





from django.views.decorators.http import require_http_methods


@require_http_methods(["DELETE"])
def delete_report(request):
    date_str = request.GET.get("date")
    if not date_str:
        return JsonResponse({"error": "Date parameter is required"}, status=400)

    try:
        date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        start_datetime = timezone.make_aware(
            datetime.datetime.combine(date, datetime.time.min)
        )
        end_datetime = timezone.make_aware(
            datetime.datetime.combine(date, datetime.time.max)
        )

        deleted_count, _ = Report.objects.filter(
            date__range=(start_datetime, end_datetime)
        ).delete()

        if deleted_count > 0:
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Deleted {deleted_count} report entries for {date_str}",
                }
            )
        else:
            return JsonResponse(
                {"success": False, "error": f"No report entries found for {date_str}"},
                status=404,
            )

    except ValueError:
        return JsonResponse({"error": "Invalid date format"}, status=400)
    except Exception as e:
        logger.error(f"Error deleting report for {date_str}: {str(e)}")
        return JsonResponse(
            {"error": "An unexpected error occurred while deleting the report"},
            status=500,
        )


from celery.result import AsyncResult
from django.http import JsonResponse


def cancel_report_task(request):
    task_id = request.GET.get("task_id")
    if not task_id:
        return JsonResponse({"error": "Task ID is required"}, status=400)

    try:
        task = AsyncResult(task_id)
        task.revoke(terminate=True)
        return JsonResponse(
            {"success": True, "message": f"Task {task_id} canceled successfully"}
        )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@require_GET
def download_report_excel(request):
    date_str = request.GET.get("date")

    if not date_str:
        return HttpResponse("Date parameter is required", status=400)

    try:
        date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        start_datetime = timezone.make_aware(
            datetime.datetime.combine(date, datetime.time.min)
        )
        end_datetime = timezone.make_aware(
            datetime.datetime.combine(date, datetime.time.max)
        )
        reports = Report.objects.filter(date__range=(start_datetime, end_datetime))
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report {date_str}"

    headers = ["Item ID", "Product Name", "Operation", "Date", "Changes"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row, report in enumerate(reports, start=2):
        ws.cell(row=row, column=1, value=report.item_id)
        ws.cell(row=row, column=2, value=report.product_name)
        ws.cell(row=row, column=3, value=report.get_operation_display())
        ws.cell(row=row, column=4, value=report.date.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=5, value=report.changes)

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=report_{date_str}.xlsx"

    return response


# @require_GET
# def get_generate_report_progress(request):
#     progress_info = cache.get(
#         "generate_report_progress", {"progress": 0, "completed": False}
#     )
#     return JsonResponse(progress_info)


import datetime
import logging

from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Max
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET

from .models import ProductChangeLog

logger = logging.getLogger(__name__)


import datetime
import io
import logging
import time
from datetime import datetime
from datetime import time as time_module

from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Max
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET
from openpyxl import Workbook

from .models import ProductChangeLog

logger = logging.getLogger(__name__)


@require_GET
def fetch_changelog(request):
    try:
        # Get date from request, default to the latest date if not provided
        date_str = request.GET.get("date")

        if date_str:
            try:
                selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({"error": "Invalid date format"}, status=400)
        else:
            latest_log_date = ProductChangeLog.objects.aggregate(Max("date"))[
                "date__max"
            ]
            if latest_log_date is None:
                return JsonResponse({"error": "No changelog entries found"}, status=404)
            selected_date = latest_log_date.date()

        start_date = timezone.make_aware(
            datetime.combine(selected_date, time_module.min)
        )
        end_date = timezone.make_aware(
            datetime.combine(selected_date, time_module.max)
        )

        logs = ProductChangeLog.objects.filter(
            date__range=(start_date, end_date)
        ).order_by("-date")

        logger.info(f"Fetching all changelog entries for {selected_date}")

        data = []
        for log in logs:
            try:
                log_data = {
                    "item_id": log.item_id,
                    "product_name": log.product_name,
                    "operation": log.get_operation_display(),
                    "date": log.date.isoformat(),
                    "changes": log.changes,
                }
                data.append(log_data)
            except Exception as e:
                logger.error(f"Error processing log entry {log.id}: {str(e)}")

        return JsonResponse(
            {
                "date": selected_date.isoformat(),
                "data": data,
            },
            safe=False,
            encoder=DjangoJSONEncoder,
        )

    except Exception as e:
        logger.error(f"Unexpected error in fetch_changelog: {str(e)}", exc_info=True)
        return JsonResponse({"error": "An unexpected error occurred"}, status=500)


from django.views.decorators.http import require_GET
from openpyxl import Workbook


@require_GET
def download_excel(request):
    date_str = request.GET.get("date")

    if not date_str:
        return HttpResponse("Date parameter is required", status=400)

    try:
        date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        start_datetime = timezone.make_aware(
            datetime.datetime.combine(date, datetime.time.min)
        )
        end_datetime = timezone.make_aware(
            datetime.datetime.combine(date, datetime.time.max)
        )
        logs = ProductChangeLog.objects.filter(
            date__range=(start_datetime, end_datetime)
        )
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Changelog {date_str}"

    # Write headers
    headers = ["Item ID", "Product Name", "Operation", "Date"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, log in enumerate(logs, start=2):
        ws.cell(row=row, column=1, value=log.item_id)
        ws.cell(row=row, column=2, value=log.product_name)
        ws.cell(row=row, column=3, value=log.get_operation_display())
        ws.cell(row=row, column=4, value=log.date.strftime("%Y-%m-%d %H:%M:%S"))

    # Save the workbook to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Prepare the response
    response = HttpResponse(
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=changelog_{date_str}.xlsx"

    return response


# Cart


import json
import uuid
from decimal import Decimal

from django.conf import settings
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from square.client import Client

from .models import Cart, Order, OrderItem, ShippingAddress


@csrf_exempt
@transaction.atomic
def process_payment(request):
    print("Starting payment processing...")

    def calculate_usps_media_mail_cost(weight):
        rate_brackets = [
            (1, 4.63), (2, 5.37), (3, 6.11), (4, 6.85), (5, 7.59),
            (6, 8.33), (7, 9.07), (8, 9.82), (9, 10.57), (10, 11.32),
            (15, 15.07), (20, 18.82), (30, 26.32), (40, 33.82),
            (50, 41.32), (60, 48.82), (70, 56.32),
        ]
        rounded_weight = int(weight) if weight == int(weight) else int(weight) + 1
        for max_weight, rate in rate_brackets:
            if rounded_weight <= max_weight:
                return rate
        return None

    def create_square_customer(client, shipping_data):
        try:
            # Format phone number to E.164 format
            phone = shipping_data['phone'].replace('(', '').replace(')', '').replace(' ', '').replace('-', '')
            if phone.startswith('0'):  # Remove leading zero if present
                phone = phone[1:]
            phone = f'+1{phone}' if not phone.startswith('+') else phone

            customer_body = {
                'given_name': shipping_data['first_name'],
                'family_name': shipping_data['last_name'],
                'email_address': shipping_data['email'],
                'address': {
                    'address_line_1': shipping_data['address_line1'],
                    'address_line_2': shipping_data['address_line2'],
                    'locality': shipping_data['city'],
                    'administrative_district_level_1': shipping_data['state'],
                    'postal_code': shipping_data['postal_code'],
                    'country': 'US',
                },
                'phone_number': phone
            }
            print("Creating Square customer with body:", json.dumps(customer_body, indent=2))
            result = client.customers.create_customer(body=customer_body)
            if result.is_success():
                return result.body.get('customer', {}).get('id')
            else:
                print("Failed to create Square customer:", result.errors)
                raise ValueError(f"Failed to create Square customer: {result.errors}")
        except Exception as e:
            print(f"Error creating Square customer: {e}")
            raise

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)

    try:
        # Start transaction savepoint
        sid = transaction.savepoint()

        data = json.loads(request.body)
        source_id = data.get('sourceId', '').strip()
        if not source_id:
            return JsonResponse({
                'error': 'Invalid payment source',
                'details': 'Source ID is required'
            }, status=400)

        cart_id = request.session.get('cart_id')
        shipping_data = request.session.get('shipping_data')
        if not cart_id or not shipping_data:
            raise ValueError("Missing cart or shipping information")

        cart = Cart.objects.get(id=cart_id)
        
        # Calculate totals
        cart_total = cart.total_amount()
        shipping_cost = calculate_usps_media_mail_cost(cart.total_weight()) + 2.00
        total_including_shipping = cart_total + Decimal(shipping_cost)

        # Initialize Square client
        client = Client(
            access_token=settings.SQUARE_ACCESS_TOKEN,
            environment='production'
        )

        # Create Square customer first
        customer_id = create_square_customer(client, shipping_data)
        if not customer_id:
            transaction.savepoint_rollback(sid)
            return JsonResponse({
                'error': 'Failed to create Square customer'
            }, status=400)

        # Create the payment body
        payment_body = {
            'source_id': source_id,
            'idempotency_key': str(uuid.uuid4()),
            'amount_money': {
                'amount': int(total_including_shipping * 100),
                'currency': 'USD'
            },
            'autocomplete': True,
            'location_id': settings.SQUARE_LOCATION_ID,
            'customer_id': customer_id,
            'line_items': [
            {
                'name': item.product.title, # Use item.product.title instead of item.product_title
                'quantity': str(item.quantity),
                'base_price_money': {
                'amount': str(item.product.price), # Keep as decimal
                'currency': 'USD'
            },
                'total_money': {
                'amount': str(item.product.price * item.quantity), # Keep as decimal
                'currency': 'USD'
            }
            } for item in cart.cartitem_set.all()
            ],
            'shipping_address': {
                'address_line_1': shipping_data['address_line1'],
                'address_line_2': shipping_data.get('address_line2', ''),
                'locality': shipping_data['city'],
                'administrative_district_level_1': shipping_data['state'],
                'postal_code': shipping_data['postal_code'],
                'country': 'US'
            }
        }

        # Process payment
        result = client.payments.create_payment(body=payment_body)
        if not result.is_success():
            transaction.savepoint_rollback(sid)
            return JsonResponse({
                'error': 'Payment processing failed',
                'details': result.errors
            }, status=400)

        # If payment successful, create order and related records
        shipping_address = ShippingAddress.objects.create(**shipping_data)
        order = Order.objects.create(
            cart=cart,
            shipping_address=shipping_address,
            subtotal=cart_total,
            shipping_cost=shipping_cost,
            total_amount=total_including_shipping,
            status='completed',
            square_payment_id=result.body.get('payment', {}).get('id')
        )

        # Create order items
        for item in cart.cartitem_set.all():
            OrderItem.objects.create(
                order=order,
                product_id=item.product.item_id,
                product_title=item.product.title,
                quantity=item.quantity,
                price=item.product.price,
            )

        # Commit the transaction
        transaction.savepoint_commit(sid)

        # Clear session data
        del request.session['cart_id']
        del request.session['shipping_data']

        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'square_payment_id': order.square_payment_id
        })

    except Exception as e:
        if 'sid' in locals():
            transaction.savepoint_rollback(sid)
        print(f"Payment processing error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': 'An error occurred processing payment',
            'details': str(e)
        }, status=500)
    

def add_to_cart(request, product_id):
    product = get_object_or_404(Product, item_id=product_id)
    cart, created = Cart.objects.get_or_create(id=request.session.get("cart_id"))

    print("Proceeding for weight checking")
    weight = get_ebay_product_weight(product.item_id)
    print(f"Fetched weight for product {product_id}: {weight}")  # Debugging output

    if weight is None:
        messages.error(
            request,
            f"Unable to fetch weight for {product.title}. Please try again later.",
        )
        return redirect("product_list")

    if not created:
        cart_item, item_created = CartItem.objects.get_or_create(
            cart=cart, product=product
        )
        if not item_created:
            cart_item.quantity += 1
        cart_item.weight = weight  # Update weight
        cart_item.save()
    else:
        request.session["cart_id"] = cart.id
        CartItem.objects.create(cart=cart, product=product, quantity=1, weight=weight)

    messages.success(request, f"{product.title} has been added to your cart.")
    return redirect("product_list")


import math
import traceback
from decimal import ROUND_HALF_UP, Decimal

from django.contrib import messages
from django.db import models
from django.db.models import Q
from django.db.models.functions import Cast
from django.shortcuts import render
from django.utils import timezone

from .models import Cart, Discount


def view_cart(request):
    try:
        print("Starting to process the cart...")

        cart_id = request.session.get("cart_id")
        cart_count = 0
        cart_total = Decimal('0.00')
        cart_items = []
        applicable_discounts = []
        total_product_discount = Decimal('0.00')
        total_cart_discount = Decimal('0.00')
        voucher_discount = Decimal('0.00')

        context = {
            "cart_items": cart_items,
            "cart_total": cart_total,
            "cart_count": cart_count,
            "applicable_discounts": applicable_discounts,
            "total_product_discount": total_product_discount,
            "total_cart_discount": total_cart_discount,
            "voucher_discount": voucher_discount,
        }

        if cart_id:
            try:
                print(f"Cart ID found: {cart_id}")
                cart = Cart.objects.get(id=cart_id)
                
                # Pre-calculate subtotals and apply discounts
                cart_items = []
                for item in cart.cartitem_set.select_related("product").all():
                    # Calculate base subtotal
                    base_subtotal = Decimal(str(item.product.price * item.quantity))
                    
                    # Track item-specific discount
                    item_product_discount = Decimal('0.00')

                    # Collect applicable non-coupon discounts
                    now = timezone.now()
                    item_applicable_discounts = Discount.objects.filter(
                        Q(is_active=True) &
                        Q(start_date__lte=now) &
                        (Q(end_date__isnull=True) | Q(end_date__gte=now)) &
                        ~Q(discount_type='COUPON')  # Exclude coupon type discounts
                    )

                    # Apply product-specific discounts
                    applied_product_discounts = []
                    for discount in item_applicable_discounts:
                        # Discount for all product prices
                        if discount.apply_to == 'PRODUCT_PRICE':
                            if discount.discount_type == 'PERCENTAGE':
                                # Calculate percentage discount
                                discount_amount = Decimal(
                                    str(item.product.price * item.quantity * (discount.discount_value / 100))
                                )
                                # Ensure discount doesn't make price negative
                                discount_amount = min(discount_amount, base_subtotal)
                                item_product_discount += discount_amount
                                applied_product_discounts.append({
                                    'name': discount.name,
                                    'amount': discount_amount,
                                    'type': 'PRODUCT_PRICE',
                                    'description': f"{discount.discount_value}% off"
                                })
                            else:
                                # Calculate fixed amount discount
                                discount_amount = Decimal(str(discount.discount_value * item.quantity))
                                # Ensure discount doesn't make price negative
                                discount_amount = min(discount_amount, base_subtotal)
                                item_product_discount += discount_amount
                                applied_product_discounts.append({
                                    'name': discount.name,
                                    'amount': discount_amount,
                                    'type': 'PRODUCT_PRICE',
                                    'description': f"${discount.discount_value} off"
                                })
                        
                        # Discount for specific products
                        elif discount.apply_to == 'SPECIFIC_PRODUCTS':
                            product_tags = discount.get_product_tags()
                            # Check if any of the tags match the product title
                            if any(tag.lower() in item.product.title.lower() for tag in product_tags):
                                if discount.discount_type == 'PERCENTAGE':
                                    # Calculate percentage discount
                                    discount_amount = Decimal(
                                        str(item.product.price * item.quantity * (discount.discount_value / 100))
                                    )
                                    # Ensure discount doesn't make price negative
                                    discount_amount = min(discount_amount, base_subtotal)
                                    item_product_discount += discount_amount
                                    applied_product_discounts.append({
                                        'name': discount.name,
                                        'amount': discount_amount,
                                        'type': 'SPECIFIC_PRODUCTS',
                                        'tags': product_tags,
                                        'description': f"{discount.discount_value}% off selected items"
                                    })
                                else:
                                    # Calculate fixed amount discount
                                    discount_amount = Decimal(str(discount.discount_value * item.quantity))
                                    # Ensure discount doesn't make price negative
                                    discount_amount = min(discount_amount, base_subtotal)
                                    item_product_discount += discount_amount
                                    applied_product_discounts.append({
                                        'name': discount.name,
                                        'amount': discount_amount,
                                        'type': 'SPECIFIC_PRODUCTS',
                                        'tags': product_tags,
                                        'description': f"${discount.discount_value} off selected items"
                                    })

                    # Calculate final item subtotal after product discounts
                    item_subtotal = max(base_subtotal - item_product_discount, Decimal('0.00'))
                    
                    # Create an enhanced cart item with additional calculation details
                    enhanced_item = {
                        'id': item.id,
                        'product': item.product,
                        'quantity': item.quantity,
                        'base_subtotal': base_subtotal,
                        'product_discount': item_product_discount,
                        'applied_product_discounts': applied_product_discounts,
                        'final_subtotal': item_subtotal
                    }
                    
                    cart_items.append(enhanced_item)

                # Calculate overall cart totals
                product_subtotal = sum(Decimal(str(item['base_subtotal'])) for item in cart_items)
                total_product_discount = sum(Decimal(str(item['product_discount'])) for item in cart_items)

                # Calculate cart-level discounts (excluding coupons)
                total_cart_discount = Decimal('0.00')
                applied_cart_discounts = []
                for discount in item_applicable_discounts:
                    if discount.apply_to == 'CART':
                        if discount.discount_type == 'PERCENTAGE':
                            cart_discount = Decimal(str(product_subtotal * (discount.discount_value / 100)))
                            total_cart_discount += cart_discount
                            applied_cart_discounts.append({
                                'name': discount.name,
                                'amount': cart_discount,
                                'description': f"{discount.discount_value}% off entire cart"
                            })
                        else:
                            cart_discount = min(Decimal(str(discount.discount_value)), product_subtotal)
                            total_cart_discount += cart_discount
                            applied_cart_discounts.append({
                                'name': discount.name,
                                'amount': cart_discount,
                                'description': f"${discount.discount_value} off entire cart"
                            })

                # Calculate intermediate cart total before voucher
                cart_total = product_subtotal - total_product_discount - total_cart_discount
                cart_total = max(cart_total, Decimal('0.00'))

                # Handle voucher application
                applied_voucher = request.session.get('applied_voucher')
                if applied_voucher:
                    voucher_code = applied_voucher.get('code')
                    try:
                        voucher = Discount.objects.get(
                            name__iexact=voucher_code,
                            discount_type='COUPON',
                            is_active=True
                        )
                        
                        if voucher.is_valid():
                            subtotal_before_voucher = cart_total
                            
                            # Determine discount type and format description accordingly
                            is_percentage = voucher.description.lower() == 'percentage discount'
                            discount_description = (
                                f"{voucher.discount_value}% off" if is_percentage
                                else f"${voucher.discount_value} off"
                            )
                            
                            # Apply voucher discount based on type and description
                            if voucher.apply_to == 'CART':
                                if is_percentage:
                                    voucher_discount = subtotal_before_voucher * (Decimal(str(voucher.discount_value)) / 100)
                                else:
                                    voucher_discount = min(Decimal(str(voucher.discount_value)), subtotal_before_voucher)
                                    
                            elif voucher.apply_to == 'PRODUCT_PRICE':
                                if is_percentage:
                                    voucher_discount = sum(
                                        item['final_subtotal'] * (Decimal(str(voucher.discount_value)) / 100)
                                        for item in cart_items
                                    )
                                else:
                                    voucher_discount = sum(
                                        min(Decimal(str(voucher.discount_value)) * item['quantity'], item['final_subtotal'])
                                        for item in cart_items
                                    )
                                    
                            elif voucher.apply_to == 'SPECIFIC_PRODUCTS':
                                product_tags = voucher.get_product_tags()
                                voucher_discount = Decimal('0.00')
                                for item in cart_items:
                                    if any(tag.lower() in item['product'].title.lower() for tag in product_tags):
                                        if is_percentage:
                                            item_discount = item['final_subtotal'] * (Decimal(str(voucher.discount_value)) / 100)
                                        else:
                                            item_discount = min(Decimal(str(voucher.discount_value)) * item['quantity'], item['final_subtotal'])
                                        voucher_discount += item_discount

                                # Update description for specific products
                                discount_description = f"{discount_description} on selected items"

                            # Ensure voucher discount doesn't make total negative
                            voucher_discount = min(voucher_discount, subtotal_before_voucher)
                            cart_total = max(subtotal_before_voucher - voucher_discount, Decimal('0.00'))
                            
                            # Create detailed voucher information
                            context['voucher_applied'] = {
                                'code': voucher_code,
                                'amount': voucher_discount,
                                'description': f"{voucher.name}: {discount_description}",
                                'type': 'percentage' if is_percentage else 'fixed',
                                'value': voucher.discount_value,
                                'apply_to': voucher.apply_to
                            }
                            
                    except Discount.DoesNotExist:
                        del request.session['applied_voucher']
                        messages.error(request, "The previously applied voucher is no longer valid.")

                # Finalize cart total and round to 2 decimal places
                cart_total = cart_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                # Update final context
                context.update({
                    "cart_items": cart_items,
                    "cart_total": cart_total,
                    "cart_count": len(cart_items),
                    "applicable_discounts": item_applicable_discounts,
                    "total_product_discount": total_product_discount,
                    "total_cart_discount": total_cart_discount,
                    "voucher_discount": voucher_discount,
                    "applied_product_discounts": applied_product_discounts,
                    "applied_cart_discounts": applied_cart_discounts
                })

                # Update session with cart details
                request.session['cart_total'] = float(cart_total)
                request.session['discounted_cart_total'] = float(cart_total)
                request.session['total_product_discount'] = float(total_product_discount)
                request.session['total_cart_discount'] = float(total_cart_discount)
                request.session['voucher_discount'] = float(voucher_discount)

                # Logging for debugging
                print(f"Final Cart Total: {cart_total}")
                print(f"Total Product Discount: {total_product_discount}")
                print(f"Total Cart Discount: {total_cart_discount}")
                print(f"Voucher Discount: {voucher_discount}")

            except Cart.DoesNotExist:
                print("Cart not found in database")
                del request.session['cart_id']
            except Exception as e:
                print("Error processing cart:")
                print(traceback.format_exc())
                messages.error(request, f"An error occurred while processing your cart: {str(e)}")

        return render(request, "pages/cart.html", context)

    except Exception as e:
        print("Unexpected error in view_cart:")
        print(traceback.format_exc())
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return render(request, "pages/cart.html", context)

def update_cart(request, item_id):
    try:
        cart_item = get_object_or_404(CartItem, id=item_id)
        if request.method == "POST":
            quantity = int(request.POST.get("quantity", 1))
            if quantity > 0:
                cart_item.quantity = quantity
                cart_item.save()
            else:
                cart_item.delete()
        return redirect("view_cart")
    except ValueError:
        messages.error(request, "Invalid quantity value")
        return redirect("view_cart")


def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id)
    cart_item.delete()
    return redirect("view_cart")


from square.client import Client

logger = logging.getLogger(__name__)


import json
import uuid
from decimal import ROUND_CEILING, ROUND_HALF_UP, Decimal

from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from square.client import Client

from .forms import ShippingAddressForm
from .models import Cart, Discount, Order, OrderItem, ShippingAddress


def checkout(request):
    try:
        print("Starting checkout process...")
        
        # Handle AJAX POST request for shipping form
        if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                data = json.loads(request.body)
                form = ShippingAddressForm(data)
                if form.is_valid():
                    request.session['shipping_data'] = form.cleaned_data
                    return JsonResponse({'success': True})
                else:
                    return JsonResponse({
                        'success': False,
                        'errors': form.errors
                    })
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'errors': {'form': ['Invalid form data']}
                }, status=400)
        
        # Get cart
        cart_id = request.session.get("cart_id")
        print(f"Cart ID from session: {cart_id}")
        
        if not cart_id:
            print("No cart ID found in session")
            return redirect("view_cart")
        
        cart = get_object_or_404(Cart, id=cart_id)
        cart_items = cart.cartitem_set.all()
        print(f"Found cart with {cart_items.count()} items")
        
        if not cart_items.exists():
            print("Cart is empty")
            messages.error(request, "Your cart is empty")
            return redirect("view_cart")
        
        # Calculate shipping cost function
        def calculate_usps_media_mail_cost(weight):
            rate_brackets = [
                (1, 4.63),
                (2, 5.37),
                (3, 6.11),
                (4, 6.85),
                (5, 7.59),
                (6, 8.33),
                (7, 9.07),
                (8, 9.82),
                (9, 10.57),
                (10, 11.32),
                (15, 15.07),
                (20, 18.82),
                (30, 26.32),
                (40, 33.82),
                (50, 41.32),
                (60, 48.82),
                (70, 56.32),
            ]

            rounded_weight = int(weight) if weight == int(weight) else int(weight) + 1
            
            for max_weight, rate in rate_brackets:
                if rounded_weight <= max_weight:
                    return Decimal(str(rate))
                    
            return None

        now = timezone.now()
        
        # Get voucher information safely
        applied_voucher = request.session.get('applied_voucher', {})
        voucher_code = applied_voucher.get('code', '').strip() if applied_voucher else ''
        
        # Get applicable discounts including coupons
        applicable_discounts = Discount.objects.filter(
            Q(is_active=True) &
            Q(start_date__lte=now) &
            (Q(end_date__isnull=True) | Q(end_date__gte=now))
        )

        detailed_cart_items = []
        cart_subtotal = Decimal('0.00')
        total_product_discount = Decimal('0.00')

        for item in cart_items:
            original_subtotal = item.product.price * item.quantity
            cart_subtotal += original_subtotal
            
            item_discounts = []
            total_item_discount = Decimal('0.00')
            
            for discount in applicable_discounts:
                if discount.discount_type != 'COUPON' or (
                    discount.discount_type == 'COUPON' and 
                    voucher_code and
                    discount.name.lower() == voucher_code.lower()
                ):
                    is_applicable = False
                    
                    if discount.apply_to == 'PRODUCT_PRICE':
                        is_applicable = True
                    elif discount.apply_to == 'SPECIFIC_PRODUCTS':
                        product_tags = discount.get_product_tags()
                        is_applicable = any(
                            tag.lower() in item.product.title.lower() 
                            for tag in product_tags
                        )
                    
                    if is_applicable:
                        # Handle both percentage and fixed discounts
                        if discount.discount_type == 'COUPON':
                            is_percentage = discount.description.lower() == 'percentage discount'
                        else:
                            is_percentage = discount.discount_type == 'PERCENTAGE'

                        if is_percentage:
                            discount_amount = original_subtotal * (discount.discount_value / 100)
                        else:
                            discount_amount = min(
                                discount.discount_value * item.quantity,
                                original_subtotal
                            )
                        
                        total_item_discount += discount_amount
                        item_discounts.append({
                            'name': discount.name,
                            'type': 'PERCENTAGE' if is_percentage else 'FIXED',
                            'value': f"{discount.discount_value}%" if is_percentage else f"${discount.discount_value}",
                            'amount': discount_amount
                        })
            
            final_subtotal = max(original_subtotal - total_item_discount, Decimal('0.00'))
            per_unit_price = final_subtotal / item.quantity if item.quantity > 0 else Decimal('0.00')
            
            detailed_cart_items.append({
                'item': item,
                'original_price': item.product.price,
                'original_subtotal': original_subtotal,
                'discounts': item_discounts,
                'total_discount': total_item_discount,
                'final_subtotal': final_subtotal,
                'per_unit_price': per_unit_price
            })
            
            total_product_discount += total_item_discount
        
        # Calculate cart-wide discounts
        subtotal_after_product_discounts = cart_subtotal - total_product_discount
        total_cart_discount = Decimal('0.00')
        applied_cart_discounts = []
        voucher_discount = Decimal('0.00')
        
        for discount in applicable_discounts:
            if discount.apply_to == 'CART':
                if discount.discount_type != 'COUPON' or (
                    discount.discount_type == 'COUPON' and 
                    voucher_code and
                    discount.name.lower() == voucher_code.lower()
                ):
                    # Determine if percentage or fixed discount
                    if discount.discount_type == 'COUPON':
                        is_percentage = discount.description.lower() == 'percentage discount'
                    else:
                        is_percentage = discount.discount_type == 'PERCENTAGE'

                    if is_percentage:
                        cart_discount = subtotal_after_product_discounts * (discount.discount_value / 100)
                    else:
                        cart_discount = min(
                            discount.discount_value,
                            subtotal_after_product_discounts
                        )
                    
                    if discount.discount_type == 'COUPON':
                        voucher_discount = cart_discount
                    else:
                        total_cart_discount += cart_discount
                        
                    applied_cart_discounts.append({
                        'name': discount.name,
                        'type': 'PERCENTAGE' if is_percentage else 'FIXED',
                        'value': discount.discount_value,
                        'amount': cart_discount
                    })
        
        # Calculate final totals
        cart_total = max(subtotal_after_product_discounts - total_cart_discount - voucher_discount, Decimal('0.00'))
        cart_total = cart_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # Calculate shipping and final total
        cart_total_weight = cart.total_weight()
        total_weight_major = int(cart_total_weight)
        total_weight_minor = int((cart_total_weight - total_weight_major) * 16)
        
        shipping_cost = calculate_usps_media_mail_cost(cart_total_weight) + Decimal('2.00')
        total_including_shipping = cart_total + shipping_cost
        total_including_shipping = total_including_shipping.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        request.session['total_including_shipping'] = str(total_including_shipping)
        request.session['cart_total'] = str(cart_total)
        request.session['shipping_cost'] = str(shipping_cost)
        
        context = {
            'detailed_cart_items': detailed_cart_items,
            'cart_subtotal': cart_subtotal,
            'cart_total': cart_total,
            'total_product_discount': total_product_discount,
            'total_cart_discount': total_cart_discount,
            'voucher_discount': voucher_discount,
            'applied_cart_discounts': applied_cart_discounts,
            'total_weight_major': total_weight_major,
            'total_weight_minor': total_weight_minor,
            'shipping_cost': shipping_cost,
            'total_including_shipping': total_including_shipping,
            'square_application_id': settings.SQUARE_APPLICATION_ID,
            'square_location_id': settings.SQUARE_LOCATION_ID,
            'cart_count': cart_items.count(),
        }
        
        # Only add voucher information to context if there's a valid voucher
        if voucher_code:
            context['voucher_applied'] = {
                'code': voucher_code,
                'amount': voucher_discount,
                'description': f"{voucher_code}: {applied_voucher.get('discount_value')}% off",
                'type': applied_voucher.get('discount_type', 'percentage'),
                'value': applied_voucher.get('discount_value'),
                'apply_to': applied_voucher.get('apply_to')
            }
        
        return render(request, "pages/checkout.html", context)
        
    except Exception as e:
        print(f"Checkout error: {str(e)}")
        import traceback
        print("Full traceback:")
        print(traceback.format_exc())
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': {'form': [str(e)]}
            }, status=500)
            
        messages.error(request, "An error occurred during checkout. Please try again.")
        return redirect("view_cart")


logger = logging.getLogger(__name__)


def order_confirmation(request, order_id):
    try:
        order = get_object_or_404(Order, id=order_id)
        order_items = order.cart.cartitem_set.all()
        order_total = order.total_amount

        return render(
            request,
            "pages/order_confirmation.html",
            {"order": order, "order_items": order_items, "order_total": order_total},
        )

    except Exception as e:
        logger.error(
            f"Error in order_confirmation view for order_id {order_id}: {e}",
            exc_info=True,
        )
        return render(
            request, "pages/error.html", {"error": "An unexpected error occurred."}
        )


def product_detail(request, product_slug):
    product = get_object_or_404(Product, html_link=product_slug)
    html_file_path = os.path.join(
        "products", "viewproduct", f"{product.html_link}.html"
    )

    if os.path.exists(html_file_path):
        with open(html_file_path, "r") as file:
            html_content = file.read()
        return HttpResponse(html_content)
    else:
        return HttpResponseNotFound("Page not found")


def get_ebay_product_weight(item_id):
    print(f"Fetching weight for item ID: {item_id}")
    url = "https://api.ebay.com/ws/api.dll"
    headers = {
        "X-EBAY-API-SITEID": "0",
        "X-EBAY-API-COMPATIBILITY-LEVEL": "967",
        "X-EBAY-API-CALL-NAME": "GetItem",
        "X-EBAY-API-IAF-TOKEN": EBAY_AUTH_TOKEN,
        "Content-Type": "text/xml",
    }
    body = f"""
    <?xml version="1.0" encoding="utf-8"?>
    <GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">    
        <ErrorLanguage>en_US</ErrorLanguage>
        <WarningLevel>High</WarningLevel>
        <ItemID>{item_id}</ItemID>
    </GetItemRequest>
    """

    while True:
        try:
            response = requests.post(url, headers=headers, data=body)
            print("Response Status Code:", response.status_code)

            if response.status_code == 200:
                response_content = response.content.decode("utf-8")

                if (
                    "<ShortMessage>Expired IAF token.</ShortMessage>"
                    in response_content
                ):
                    print("Expired IAF token error detected")
                    print("Access token expired, fetching a new one...")
                    new_token = fetch_new_access_token()
                    if new_token:
                        headers["X-EBAY-API-IAF-TOKEN"] = new_token
                        print("New token fetched and set")
                        continue
                    else:
                        print("Failed to fetch a new access token")
                        return None

                root = ET.fromstring(response_content)

                namespace = {"ns": "urn:ebay:apis:eBLBaseComponents"}

                weight_major_elem = root.find(
                    ".//ns:ShippingPackageDetails/ns:WeightMajor", namespace
                )
                weight_minor_elem = root.find(
                    ".//ns:ShippingPackageDetails/ns:WeightMinor", namespace
                )

                if weight_major_elem is not None:
                    weight_major = weight_major_elem.text
                else:
                    print("WeightMajor element not found")
                    weight_major = "0"

                if weight_minor_elem is not None:
                    weight_minor = weight_minor_elem.text
                else:
                    print("WeightMinor element not found")
                    weight_minor = "0"

                weight_major = float(weight_major)
                weight_minor = float(weight_minor)
                print(f"{item_id} - weight - {weight_major} , {weight_minor}")

                total_weight_lbs = weight_major + (weight_minor / 16)
                return total_weight_lbs

            else:
                print("Failed to fetch data from eBay API")
                print("Response Content:", response_content)
                return None

        except Exception as e:
            print("Exception occurred:", str(e))
            return None


def get_ebay_listing_status(item_id):
    print(f"Fetching status for item ID: {item_id}")
    url = "https://api.ebay.com/ws/api.dll"
    headers = {
        "X-EBAY-API-SITEID": "0",
        "X-EBAY-API-COMPATIBILITY-LEVEL": "967",
        "X-EBAY-API-CALL-NAME": "GetItem",
        "X-EBAY-API-IAF-TOKEN": EBAY_AUTH_TOKEN,
        "Content-Type": "text/xml",
    }
    body = f"""
    <?xml version="1.0" encoding="utf-8"?>
    <GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">    
        <ErrorLanguage>en_US</ErrorLanguage>
        <WarningLevel>High</WarningLevel>
        <ItemID>{item_id}</ItemID>
    </GetItemRequest>
    """

    max_retries = 3
    retry_delay = 5  # seconds

    for attempt in range(max_retries):
        try:
            response = requests.post(url, headers=headers, data=body)
            print("Response Status Code:", response.status_code)

            if response.status_code == 200:
                response_content = response.content.decode("utf-8")

                if (
                    "<ShortMessage>Expired IAF token.</ShortMessage>"
                    in response_content
                ):
                    print("Expired IAF token error detected")
                    print("Access token expired, fetching a new one...")
                    new_token = fetch_new_access_token()
                    if new_token:
                        headers["X-EBAY-API-IAF-TOKEN"] = new_token
                        print("New token fetched and set")
                        continue
                    else:
                        print("Failed to fetch a new access token")
                        return None

                root = ET.fromstring(response_content)

                namespace = {"ns": "urn:ebay:apis:eBLBaseComponents"}

                status_elem = root.find(".//ns:ListingStatus", namespace)

                if status_elem is not None:
                    listing_status = status_elem.text
                    print(f"{item_id} - status - {listing_status}")
                    return listing_status
                else:
                    print("ListingStatus element not found")
                    return None

            else:
                print("Failed to fetch data from eBay API")
                print("Response Content:", response.content.decode("utf-8"))

                if attempt < max_retries - 1:
                    print(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                else:
                    print("Max retries reached. Giving up.")
                    return None

        except Exception as e:
            print("Exception occurred:", str(e))
            if attempt < max_retries - 1:
                print(f"Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
            else:
                print("Max retries reached. Giving up.")
                return None

    return None  # If we've exhausted all retries


# Calendar
import json
from datetime import datetime

from django.contrib import messages
from django.http import JsonResponse
from django.middleware.csrf import get_token
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import CalendarEvent


def add_event(request):
    if request.method == "POST":
        try:
            # Check if eventId exists to update an event, otherwise create a new one
            event_id = request.POST.get("eventId")
            if event_id:
                event = get_object_or_404(CalendarEvent, id=event_id)
            else:
                event = CalendarEvent()

            # Setting event data from form input
            event.title = request.POST["title"]
            event.description = request.POST.get("description", "")

            # Parsing and setting start and end dates
            start_date = datetime.strptime(request.POST["start_date"], "%Y-%m-%dT%H:%M")
            end_date = datetime.strptime(request.POST["end_date"], "%Y-%m-%dT%H:%M")

            # Storing as timezone-aware datetimes
            event.start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
            event.end_date = timezone.make_aware(end_date, timezone.get_current_timezone())

            # Storing location
            event.location = request.POST.get("location", "")
            event.save()

            # Success message and return JSON response
            return JsonResponse({
                "message": "Event saved successfully!", 
                "event": {
                    "id": event.id,
                    "title": event.title,
                    "start_date": event.start_date.strftime("%Y-%m-%d %H:%M"),
                    "end_date": event.end_date.strftime("%Y-%m-%d %H:%M"),
                    "location": event.location
                }
            }, status=200)

        except Exception as e:
            # Handle exceptions and return error response
            return JsonResponse({"message": f"Failed to save event: {str(e)}"}, status=400)

    # Fetch all events to display them below the form
    events = CalendarEvent.objects.all().order_by("-start_date")
    return render(request, "pages/admin-4.html", {"events": events})

@require_POST
def delete_event(request, event_id):
    try:
        event = get_object_or_404(CalendarEvent, id=event_id)
        event.delete()
        return JsonResponse({"message": "Event deleted successfully!"}, status=200)
    except Exception as e:
        return JsonResponse({"message": f"Failed to delete event: {str(e)}"}, status=400)

def admin_page4(request):
    events = CalendarEvent.objects.all().order_by("-start_date")
    return render(request, "pages/admin-4.html", {"events": events})



# Discount
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy

from .forms import DiscountForm
from .models import Discount


def discount_list(request):
    discounts = Discount.objects.all()
    return render(request, 'pages/discount_list.html', {'discounts': discounts})

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy

from .forms import DiscountForm
from .models import Discount


# views.py
def discount_create(request):
    if request.method == 'POST':
        form = DiscountForm(request.POST)
        try:
            if form.is_valid():
                form.save()
                return redirect('discount_list')
            else:
                # More detailed error printing
                print("Form is NOT valid")
                print("Form Errors:", form.errors)
                print("Form Data:", request.POST)
                
                # Optional: Add more context
                for field, errors in form.errors.items():
                    print(f"Field {field} has errors: {errors}")
        except Exception as e:
            print(f"Unexpected error during form save: {e}")
    else:
        form = DiscountForm()
    
    return render(request, 'pages/discount_form.html', {'form': form})


def discount_update(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    if request.method == 'POST':
        form = DiscountForm(request.POST, instance=discount)
        if form.is_valid():
            form.save()
            return redirect('discount_list')
    else:
        form = DiscountForm(instance=discount)
    return render(request, 'pages/discount_form.html', {'form': form})

def discount_delete(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    if request.method == 'POST':
        discount.delete()
        return redirect('discount_list')
    return render(request, 'pages/discount_confirm_delete.html', {'object': discount})

from django.contrib import messages
# Add this to your existing views.py
from django.shortcuts import redirect


def apply_voucher(request):
    if request.method == "POST":
        voucher_code = request.POST.get('voucher_code', '').strip().upper()
        cart_id = request.session.get('cart_id')
        
        if not cart_id:
            messages.error(request, "No active cart found.")
            return redirect('view_cart')
        
        try:
            cart = Cart.objects.get(id=cart_id)
            
            # Check if voucher exists and is valid
            try:
                discount = Discount.objects.get(
                    name__iexact=voucher_code,
                    discount_type='COUPON',
                    is_active=True
                )
                
                # Validate voucher
                if not discount.is_valid():
                    messages.error(request, "This voucher has expired.")
                    return redirect('view_cart')
                
                # Check minimum purchase amount
                cart_total = Decimal(str(request.session.get('cart_total', '0')))
                if cart_total < discount.minimum_purchase_amount:
                    messages.error(request, f"Minimum purchase amount of ${discount.minimum_purchase_amount} required.")
                    return redirect('view_cart')
                
                # Store voucher in session
                request.session['applied_voucher'] = {
                    'code': voucher_code,
                    'discount_value': float(discount.discount_value),
                    'apply_to': discount.apply_to
                }
                
                messages.success(request, f"Voucher applied successfully! {discount.discount_value}% off your order.")
                
            except Discount.DoesNotExist:
                messages.error(request, "Invalid voucher code.")
                
        except Cart.DoesNotExist:
            messages.error(request, "Cart not found.")
            
    return redirect('view_cart')


def remove_voucher(request):
    if request.method == "POST":
        if 'applied_voucher' in request.session:
            del request.session['applied_voucher']
            messages.success(request, "Voucher removed successfully.")
    return redirect('view_cart')

from django.shortcuts import redirect


def order_information_redirect(request):
    return redirect('landing_page')

from django.http import JsonResponse
from django.views.decorators.cache import cache_page
from django.views.decorators.http import require_GET


@require_GET
@cache_page(60 * 60)  # Cache for 1 hour
def get_shipping_cost(request, item_id):
    def calculate_usps_media_mail_cost(weight):
        rate_brackets = [
        (1, 4.63),
        (2, 5.37),
        (3, 6.11),
        (4, 6.85),
        (5, 7.59),
        (6, 8.33),
        (7, 9.07),
        (8, 9.82),
        (9, 10.57),
        (10, 11.32),
        (15, 15.07),
        (20, 18.82),
        (30, 26.32),
        (40, 33.82),
        (50, 41.32),
        (60, 48.82),
        (70, 56.32),
    ]

    # Round up weight to the nearest whole number
        rounded_weight = int(weight) if weight == int(weight) else int(weight) + 1

        # Find the appropriate rate bracket
        for max_weight, rate in rate_brackets:
            if rounded_weight <= max_weight:
                return rate

        # If weight exceeds the highest bracket, return None or handle as needed
        return None
    try:
        # First, try to get the weight from cache to avoid repeated API calls
        cached_weight = cache.get(f'product_weight_{item_id}')
        
        if not cached_weight:
            # Fetch weight from eBay API
            cached_weight = get_ebay_product_weight(item_id)
            
            if cached_weight:
                # Cache the weight for future requests
                cache.set(f'product_weight_{item_id}', cached_weight, 24 * 60 * 60)  # Cache for 24 hours
        
        if cached_weight is not None:
            shipping_cost = calculate_usps_media_mail_cost(cached_weight)
            
            return JsonResponse({
                'shipping_cost': shipping_cost,
                'weight': cached_weight
            })
        
        return JsonResponse({'error': 'Could not fetch shipping cost'}, status=404)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


import json
import uuid
from decimal import Decimal

import requests
from django.conf import settings
from django.db import transaction
# paypal
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

# Single product update
# from django.shortcuts import render
# from django.views.generic import TemplateView

# from .models import Product, ProductChangeLog


# class ProductLookupView(TemplateView):
#     template_name = 'product_lookup.html'

#     def get(self, request, *args, **kwargs):
#         print(f"[DEBUG] ProductLookupView GET request received")
#         return render(request, self.template_name)

#     def post(self, request, *args, **kwargs):
#         print(f"[DEBUG] ProductLookupView POST request received with data: {request.POST}")
#         item_ids = request.POST.get('item_ids', '').strip()
#         item_list = [id.strip() for id in item_ids.split(',') if id.strip()]
        
#         if not item_list:
#             print("[DEBUG] No valid item IDs provided")
#             return JsonResponse({'error': 'No valid item IDs provided'})

#         results = []
#         for item_id in item_list:
#             print(f"[DEBUG] Looking up product with item_id: {item_id}")
#             product = Product.objects.filter(item_id=item_id).first()
#             results.append({
#                 'item_id': item_id,
#                 'exists': product is not None,
#                 'title': product.title if product else None
#             })
        
#         print(f"[DEBUG] Lookup results: {results}")
#         return JsonResponse({'results': results})

# def update_products(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Invalid request method'})

#     print("[DEBUG] Starting update_products function")
#     item_ids = request.POST.get('item_ids', '').strip()
#     item_list = [id.strip() for id in item_ids.split(',') if id.strip()]
#     print(f"[DEBUG] Processing items: {item_list}")
    
#     if not item_list:
#         return JsonResponse({'error': 'No valid item IDs provided'})

#     fields_to_check = ["title", "global_id", "category_id", "category_name",
#                       "gallery_url", "view_item_url", "auto_pay", "postal_code",
#                       "location", "country", "shipping_type", "ship_to_locations", 
#                       "price", "short_description"]

#     updated_items = []
#     deleted_items = []
#     errors = []

#     for item_id in item_list:
#         print(f"\n[DEBUG] Processing item_id: {item_id}")
#         try:
#             browse_data = fetch_browse_api_data(item_id)
#             # print(f"[DEBUG] Browse API data received: {browse_data}")
            
#             if not browse_data:
#                 print(f"[ERROR] No Browse API data for item_id: {item_id}")
#                 errors.append({'item_id': item_id, 'error': 'No Browse API data found'})
#                 continue

#             try:
#                 product = Product.objects.get(item_id=item_id)
#                 print(f"[DEBUG] Found product in database: {product.title}")
                
#                 # Handle deletion case first and skip the rest
#                 if 'end_time' in browse_data:
#                     print(f"[DEBUG] End time found: {browse_data['end_time']}")
#                     try:
#                         ProductChangeLog.objects.create(
#                             item_id=item_id,
#                             product_name=product.title,
#                             operation="deleted",
#                             changes={
#                                 'reason': 'Product listing ended',
#                                 'end_time': browse_data['end_time']
#                             }
#                         )
#                         product.delete()
#                         deleted_items.append({
#                             'item_id': item_id,
#                             'end_time': browse_data['end_time']
#                         })
#                         continue
#                     except Exception as e:
#                         print(f"[ERROR] Failed to delete product: {str(e)}")
#                         errors.append({'item_id': item_id, 'error': f'Deletion failed: {str(e)}'})
#                         continue

#                 # Only proceed with update logic if item is not deleted
#                 if 'price' not in browse_data:
#                     print(f"[ERROR] No price in browse_data for item_id: {item_id}")
#                     errors.append({'item_id': item_id, 'error': 'No price found in Browse API data'})
#                     continue

#                 # Rest of the update logic remains the same
#                 price = float(browse_data['price'])
#                 min_price = max(price - 0.1, 0)
#                 max_price = (price + 0.1)

#                 finding_data = None
#                 current_page = 1
                
#                 while True:
#                     print(f"[DEBUG] Fetching page {current_page} for price range: ${min_price:.2f} - ${max_price:.2f}")
#                     items, total_pages = fetch_finding_api_data(
#                         page_number=current_page,
#                         min_price=min_price,
#                         max_price=max_price
#                     )

#                     finding_data = next((item for item in items if item['item_id'] == item_id), None)
#                     if finding_data:
#                         break
                    
#                     if current_page >= total_pages:
#                         break

#                     current_page += 1

#                 if not finding_data:
#                     print(f"[ERROR] Item not found in Finding API: {item_id}")
#                     errors.append({'item_id': item_id, 'error': 'Item not found in Finding API'})
#                     continue

#                 print("[DEBUG] Preparing data for update")
#                 if 'description' in finding_data:
#                     finding_data['short_description'] = clean_description(finding_data.pop('description'))
#                 if 'description' in browse_data:
#                     browse_data['short_description'] = clean_description(browse_data.pop('description'))
#                 browse_data['price'] = finding_data['price']
#                 combined_data = {**finding_data, **browse_data}
                
#                 before_dict = {
#                     key: float(getattr(product, key)) if isinstance(getattr(product, key), Decimal) else getattr(product, key)
#                     for key in fields_to_check if hasattr(product, key)
#                 }

#                 changes = {}
#                 for key, value in combined_data.items():
#                     if key in fields_to_check and hasattr(product, key):
#                         if before_dict.get(key) != value:
#                             setattr(product, key, value)
#                             changes[key] = value

#                 if changes:
#                     try:
#                         product.save()
#                         change_log = ProductChangeLog.objects.create(
#                             item_id=item_id,
#                             product_name=product.title,
#                             operation="updated"
#                         )
#                         change_log.set_changes(before_dict, {**before_dict, **changes})
#                         change_log.save()
#                         updated_items.append({
#                             'item_id': item_id,
#                             'changes': changes
#                         })
#                     except Exception as e:
#                         print(f"[ERROR] Failed to save changes: {str(e)}")
#                         errors.append({'item_id': item_id, 'error': f'Save failed: {str(e)}'})

#             except Product.DoesNotExist:
#                 print(f"[ERROR] Product not found in database: {item_id}")
#                 errors.append({'item_id': item_id, 'error': 'Product not found in database'})

#         except Exception as e:
#             print(f"[ERROR] Unexpected error for item_id {item_id}: {str(e)}")
#             errors.append({'item_id': item_id, 'error': str(e)})
#     generate_html_pages_async.delay()
#     return JsonResponse({
#     'success': True,
#     'message': 'Products updated successfully',
#     'data': {
#         'updated': updated_items,
#         'deleted': deleted_items,
#         'errors': errors
#     }
# })

# def create_products(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Invalid request method'})

#     print("[DEBUG] Starting create_products function")
#     item_ids = request.POST.get('item_ids', '').strip()
#     item_list = [id.strip() for id in item_ids.split(',') if id.strip()]
#     print(f"[DEBUG] Processing {len(item_list)} items: {item_list}")
    
#     if not item_list:
#         return JsonResponse({'error': 'No valid item IDs provided'})

#     created_items = []
#     errors = []

#     for item_id in item_list:
#         try:
#             print(f"\n[DEBUG] Processing item_id: {item_id}")
#             browse_data = fetch_browse_api_data(item_id)
#             if not browse_data or 'price' not in browse_data:
#                 print(f"[ERROR] No Browse API data found for item_id: {item_id}")
#                 errors.append({'item_id': item_id, 'error': 'No Browse API data found'})
#                 continue

#             # Check if end_time exists and skip creation if it does
#             if 'end_time' in browse_data:
#                 print(f"[DEBUG] Item {item_id} has an end_time, skipping creation")
#                 errors.append({'item_id': item_id, 'error': 'Item has expired'})
#                 continue

#             price = float(browse_data['price'])
#             min_price = max(price - 0.1, 0)
#             max_price = (price + 0.1)
#             print(f"[DEBUG] Price range for Finding API: {min_price} - {max_price}")

#             finding_data = None
#             current_page = 1
#             print(f"[DEBUG] Starting Finding API pagination search for item_id: {item_id}")

#             while True:
#                 print(f"[DEBUG] Fetching page {current_page} for price range: ${min_price:.2f} - ${max_price:.2f}")
#                 items, total_pages = fetch_finding_api_data(
#                     page_number=current_page,
#                     min_price=min_price,
#                     max_price=max_price
#                 )
#                 print(f"[DEBUG] Retrieved page {current_page} of {total_pages} total pages")
#                 print(f"[DEBUG] Found {len(items)} items on this page")

#                 finding_data = next((item for item in items if item['item_id'] == item_id), None)
#                 if finding_data:
#                     print(f"[DEBUG] Found item {item_id} on page {current_page}")
#                     break
                
#                 if current_page >= total_pages:
#                     print(f"[DEBUG] Reached last page ({current_page}) without finding item {item_id}")
#                     break

#                 current_page += 1
#                 print(f"[DEBUG] Moving to page {current_page}")

#             if not finding_data:
#                 print(f"[ERROR] Item {item_id} not found in Finding API results")
#                 errors.append({'item_id': item_id, 'error': 'Item not found in Finding API'})
#                 continue

#             print(f"[DEBUG] Preparing data for item creation: {item_id}")
#             if 'description' in finding_data:
#                 finding_data['short_description'] = clean_description(finding_data.pop('description'))
#             if 'description' in browse_data:
#                 browse_data['short_description'] = clean_description(browse_data.pop('description'))
#             browse_data['price'] = finding_data['price']
#             combined_data = {**finding_data, **browse_data}
            
#             save_product_data(combined_data)
#             print(f"[SUCCESS] Created product for item_id: {item_id}")
#             created_items.append(item_id)
#             ProductChangeLog.objects.create(
#                 item_id=item_id,
#                 product_name=combined_data.get("title", "Unknown Title"),
#                 operation="created",
#             )

#         except Exception as e:
#             print(f"[ERROR] Exception for item_id {item_id}: {str(e)}")
#             print(f"[ERROR] Exception traceback: {traceback.format_exc()}")
#             errors.append({'item_id': item_id, 'error': str(e)})
#     generate_html_pages_async.delay()
#     print(f"[DEBUG] Final results - Created: {len(created_items)}, Errors: {len(errors)}")
#     return JsonResponse({
#     'success': True,
#     'message': 'Products created successfully',
#     'data': {
#         'created': created_items,
#         'errors': errors
#     }
# })





def get_paypal_access_token():
    try:
        # For production mode
        base_url = "https://api-m.paypal.com/v1/oauth2/token"
        
        headers = {
            "Accept": "application/json",
            "Accept-Language": "en_US",
            "Content-Type": "application/x-www-form-urlencoded"
        }
        
        # Use environment variables for credentials
        auth = (settings.PAYPAL_CLIENT_ID, settings.PAYPAL_CLIENT_SECRET)
        
        data = {
            "grant_type": "client_credentials"
        }
        
        response = requests.post(
            base_url, 
            auth=auth,
            headers=headers,
            data=data
        )
        
        if response.status_code != 200:
            raise ValueError(f"Failed to get token: {response.text}")
        print(response.json().get('access_token',''))
        return response.json().get('access_token','')
        
    except Exception as e:
        print(f"PayPal token error: {str(e)}")
        raise ValueError("Failed to authenticate with PayPal")
    

@csrf_exempt
@require_http_methods(["POST"])
def create_paypal_order(request):
    try:
        # Get cart from session
        cart_id = request.session.get("cart_id")
        if not cart_id:
            return JsonResponse({"error": "No cart found"}, status=400)
            
        # Get shipping data from request
        data = json.loads(request.body)
        shipping_data = data.get('shipping_data')
        
        if not shipping_data:
            return JsonResponse({"error": "Missing shipping information"}, status=400)
            
        # Store shipping data in session
        request.session['shipping_data'] = shipping_data
            
        cart = Cart.objects.get(id=cart_id)
        
        # Get PayPal access token
        try:
            access_token = get_paypal_access_token()
        except ValueError as e:
            return JsonResponse({"error": str(e)}, status=500)

        # Get total from session
        total_including_shipping = Decimal(request.session.get('total_including_shipping'))
        
        # Create PayPal order
        url = "https://api-m.paypal.com/v2/checkout/orders"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        payload = {
            "intent": "CAPTURE",
            "purchase_units": [{
                "amount": {
                    "currency_code": "USD",
                    "value": str(total_including_shipping)
                },
                "shipping": {
                    "name": {
                        "full_name": f"{shipping_data['first_name']} {shipping_data['last_name']}"
                    },
                    "address": {
                        "address_line_1": shipping_data['address_line1'],
                        "address_line_2": shipping_data.get('address_line2', ''),
                        "admin_area_2": shipping_data['city'],
                        "admin_area_1": shipping_data['state'],
                        "postal_code": shipping_data['postal_code'],
                        "country_code": "US"
                    }
                }
            }]
        }
        
        response = requests.post(url, headers=headers, json=payload)
        if not response.ok:
            return JsonResponse({"error": "Failed to create PayPal order"}, status=response.status_code)
            
        return JsonResponse(response.json())
        
    except Exception as e:
        print(f"Error creating PayPal order: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def capture_paypal_order(request, order_id):
    try:
        sid = transaction.savepoint()
        
        # Get cart and shipping info
        cart_id = request.session.get("cart_id")
        shipping_data = request.session.get("shipping_data")
        total_including_shipping = Decimal(request.session.get('total_including_shipping'))
        shipping_cost = Decimal(request.session.get('shipping_cost'))
        cart_total = Decimal(request.session.get('cart_total'))

        if not cart_id or not shipping_data:
            raise ValueError("Missing cart or shipping information")
            
        cart = Cart.objects.get(id=cart_id)
        
        # Get PayPal access token
        access_token = get_paypal_access_token()
        
        # Capture PayPal order
        url = f"https://api-m.paypal.com/v2/checkout/orders/{order_id}/capture"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        response = requests.post(url, headers=headers)
        capture_data = response.json()
        
        if response.status_code != 201:
            raise ValueError(f"PayPal capture failed: {capture_data.get('message')}")
            
        # Create shipping address
        shipping_address = ShippingAddress.objects.create(**shipping_data)
        
        # Create order using the total from session
        order = Order.objects.create(
            cart=cart,
            shipping_address=shipping_address,
            total_amount=total_including_shipping,
            shipping_cost = shipping_cost,
            subtotal = cart_total,
            status='completed',
            square_payment_id=f"PP_{capture_data['id']}"
        )
        
        # Create order items
        for item in cart.cartitem_set.all():
            OrderItem.objects.create(
                order=order,
                product_id=item.product.item_id,
                product_title=item.product.title,
                quantity=item.quantity,
                price=item.product.price,
            )
            
        # Clear session data
        del request.session['cart_id']
        del request.session['shipping_data']
        del request.session['total_including_shipping']
        
        # Commit transaction
        transaction.savepoint_commit(sid)
        
        return JsonResponse({
            "success": True,
            "order_id": order.id,
            "paypal_order_id": capture_data['id']
        })
        
    except Exception as e:
        if 'sid' in locals():
            transaction.savepoint_rollback(sid)
        print(f"PayPal capture error: {str(e)}")
        return JsonResponse({
            "error": "Payment processing failed",
            "details": str(e)
        }, status=500)


# single product alter
from django.http import JsonResponse
from django.shortcuts import render
from django.views.generic import TemplateView

from .models import Product, ProductChangeLog


class ProductLookupView(TemplateView):
    template_name = 'product_lookup.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        item_ids = request.POST.get('item_ids', '').strip()
        item_list = [id.strip() for id in item_ids.split(',') if id.strip()]
        
        if not item_list:
            return JsonResponse({'error': 'No valid item IDs provided'})

        results = []
        for item_id in item_list:
            product = Product.objects.filter(item_id=item_id).first()
            results.append({
                'item_id': item_id,
                'exists': product is not None,
                'title': product.title if product else None
            })
        
        return JsonResponse({'results': results})
    
def process_products(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'})

    item_ids = request.POST.get('item_ids', '').strip()
    item_list = [id.strip() for id in item_ids.split(',') if id.strip()]
    
    if not item_list:
        return JsonResponse({'error': 'No valid item IDs provided'})

    results = []
    
    for item_id in item_list:
        try:
            product_exists = Product.objects.filter(item_id=item_id).exists()
            browse_data = fetch_browse_api_data(item_id)
            
            if not browse_data:
                results.append({
                    'item_id': item_id,
                    'process': 'API Check',
                    'status': 'Failed',
                    'message': 'No Browse API data found'
                })
                continue

            if 'end_time' in browse_data:
                if product_exists:
                    result = handle_product_deletion(item_id)
                    results.append(result)
                continue

            if product_exists:
                result = update_existing_product(item_id, browse_data)
            else:
                result = create_new_product(item_id, browse_data)
            
            results.append(result)

        except Exception as e:
            results.append({
                'item_id': item_id,
                'process': 'Processing',
                'status': 'Failed',
                'message': str(e)
            })
    generate_html_pages_async.delay()
    return JsonResponse({
        'success': True,
        'results': results
    })

def handle_product_deletion(item_id):
    try:
        product = Product.objects.get(item_id=item_id)
        ProductChangeLog.objects.create(
            item_id=item_id,
            product_name=product.title,
            operation="deleted",
            changes={'reason': 'Product listing ended'}
        )
        product.delete()
        return {
            'item_id': item_id,
            'process': 'Delete',
            'status': 'Success',
            'message': 'Product deleted - listing ended'
        }
    except Exception as e:
        return {
            'item_id': item_id,
            'process': 'Delete',
            'status': 'Failed',
            'message': str(e)
        }

def update_existing_product(item_id, browse_data):
    try:
        product = Product.objects.get(item_id=item_id)
        price = float(browse_data['price'])
        finding_data = get_finding_api_data(item_id, price)
        
        if not finding_data:
            return {
                'item_id': item_id,
                'process': 'Update',
                'status': 'Failed',
                'message': 'Item not found in Finding API'
            }

        before_dict = get_product_current_state(product)
        changes = apply_product_updates(product, finding_data, browse_data)
        
        if changes:
            product.save()
            log_product_changes(item_id, product.title, before_dict, changes)
            return {
                'item_id': item_id,
                'process': 'Update',
                'status': 'Success',
                'message': f'Updated fields: {", ".join(changes.keys())}'
            }
        else:
            return {
                'item_id': item_id,
                'process': 'Update',
                'status': 'Success',
                'message': 'No changes needed'
            }
            
    except Exception as e:
        return {
            'item_id': item_id,
            'process': 'Update',
            'status': 'Failed',
            'message': str(e)
        }

def create_new_product(item_id, browse_data):
    try:
        price = float(browse_data['price'])
        finding_data = get_finding_api_data(item_id, price)
        
        if not finding_data:
            return {
                'item_id': item_id,
                'process': 'Create',
                'status': 'Failed',
                'message': 'Item not found in Finding API'
            }

        if 'description' in finding_data:
            finding_data['short_description'] = clean_description(finding_data.pop('description'))
        if 'description' in browse_data:
            browse_data['short_description'] = clean_description(browse_data.pop('description'))

        browse_data['price'] = finding_data['price']
        combined_data = {**finding_data, **browse_data}
        
        save_product_data(combined_data)
        
        ProductChangeLog.objects.create(
            item_id=item_id,
            product_name=finding_data['title'],
            operation="created"
        )
        
        return {
            'item_id': item_id,
            'process': 'Create',
            'status': 'Success',
            'message': f"Created {finding_data['title']}"
        }
        
    except Exception as e:
        return {
            'item_id': item_id,
            'process': 'Create',
            'status': 'Failed',
            'message': str(e)
        }

def get_finding_api_data(item_id, price):
    min_price = max(price - 0.1, 0)
    max_price = price + 0.1
    current_page = 1
    
    while True:
        items, total_pages = fetch_finding_api_data(
            page_number=current_page,
            min_price=min_price,
            max_price=max_price
        )
        
        finding_data = next((item for item in items if item['item_id'] == item_id), None)
        if finding_data or current_page >= total_pages:
            return finding_data
            
        current_page += 1

def get_product_current_state(product):
    fields_to_check = [
        "title", "global_id", "category_id", "category_name",
        "gallery_url", "view_item_url", "auto_pay", "postal_code",
        "location", "country", "shipping_type", "ship_to_locations", 
        "price", "short_description"
    ]
    
    return {
        key: float(getattr(product, key)) if isinstance(getattr(product, key), Decimal) 
        else getattr(product, key)
        for key in fields_to_check if hasattr(product, key)
    }

def apply_product_updates(product, finding_data, browse_data):
    if 'description' in finding_data:
        finding_data['short_description'] = clean_description(finding_data.pop('description'))
    if 'description' in browse_data:
        browse_data['short_description'] = clean_description(browse_data.pop('description'))
        
    browse_data['price'] = finding_data['price']
    combined_data = {**finding_data, **browse_data}
    
    changes = {}
    for key, value in combined_data.items():
        if hasattr(product, key) and getattr(product, key) != value:
            setattr(product, key, value)
            changes[key] = value
            
    return changes

def log_product_changes(item_id, title, before_dict, changes):
    change_log = ProductChangeLog.objects.create(
        item_id=item_id,
        product_name=title,
        operation="updated"
    )
    change_log.set_changes(before_dict, {**before_dict, **changes})
    change_log.save()


# pdf

import logging

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

logger = logging.getLogger(__name__)

def generate_order_pdf(request, order_id):
    try:
        # Get the order
        order = get_object_or_404(Order.objects.select_related('shipping_address'), id=order_id)
        
        # Create the HttpResponse object
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="order-{order.id}.pdf"'
        
        # Create the PDF document
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Styles setup
        elements = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor('#2d5c8a')
        ))
        
        # Title
        elements.append(Paragraph(f"Order #{getattr(order, 'order_number', 'N/A')}", styles['CustomHeading']))
        elements.append(Spacer(1, 10))
        
        # Order Information Table
        elements.append(Paragraph("Order Information", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        order_data = [
            ["Order Number:", str(getattr(order, 'order_number', 'N/A'))],
            ["Square Payment ID:", str(getattr(order, 'square_payment_id', 'N/A'))],
            ["Order Date:", order.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(order, 'created_at') else 'N/A'],
            ["Status:", str(getattr(order, 'get_status_display', lambda: 'N/A')())],
            ["Delivery Status:", str(getattr(order, 'get_delivery_status_display', lambda: 'N/A')())]
        ]
        
        order_table = Table(order_data, colWidths=[2*inch, 4*inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(order_table)
        elements.append(Spacer(1, 20))
        
        # Customer Information Table
        shipping_address = getattr(order, 'shipping_address', None)
        if shipping_address:
            elements.append(Paragraph("Customer Information", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            customer_data = [
                ["Name:", f"{getattr(shipping_address, 'first_name', '')} {getattr(shipping_address, 'last_name', '')}"],
                ["Email:", getattr(shipping_address, 'email', '')],
                ["Phone:", getattr(shipping_address, 'phone', '')],
                ["Address:", getattr(shipping_address, 'address_line1', '')],
            ]
            
            # Add address line 2 if it exists
            if getattr(shipping_address, 'address_line2', ''):
                customer_data.append(["", getattr(shipping_address, 'address_line2', '')])
            
            customer_data.append([
                "",
                f"{getattr(shipping_address, 'city', '')}, {getattr(shipping_address, 'state', '')} {getattr(shipping_address, 'postal_code', '')}"
            ])
            customer_data.append(["", getattr(shipping_address, 'country', '')])
            
            customer_table = Table(customer_data, colWidths=[2*inch, 4*inch])
            customer_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(customer_table)
            elements.append(Spacer(1, 20))
        
        # Order Items Table
        elements.append(Paragraph("Order Items", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        # Column widths adjusted for better product title display
        col_widths = [1.2*inch, 3.3*inch, 0.8*inch, 1*inch, 1.2*inch]
        
        items_data = [["Product ID", "Product Title", "Quantity", "Price", "Subtotal"]]
        
        # Try to get order items using different possible attribute names
        order_items = []
        possible_relations = ['items', 'order_items', 'orderitems', 'line_items']
        
        for relation in possible_relations:
            try:
                items = getattr(order, relation, None)
                if items is not None and hasattr(items, 'all'):
                    order_items = items.all()
                    break
            except Exception:
                continue

        # If we found order items, add them to the table
        if order_items:
            for item in order_items:
                # Wrap product title in Paragraph for automatic text wrapping
                product_title = Paragraph(
                    str(getattr(item, 'product_title', 'N/A')),
                    ParagraphStyle('ProductTitle', fontSize=10, leading=12)
                )
                
                items_data.append([
                    str(getattr(item, 'product_id', 'N/A')),
                    product_title,
                    str(getattr(item, 'quantity', 0)),
                    f"${getattr(item, 'price', 0):.2f}",
                    f"${getattr(item, 'subtotal', 0):.2f}"
                ])

        # Add totals with clear separation
        items_data.extend([
            ["", "", "", "", ""],  # Empty row for spacing
            ["", "", "", "Subtotal:", f"${getattr(order, 'subtotal', 0):.2f}"],
            ["", "", "", "Shipping:", f"${getattr(order, 'shipping_cost', 0):.2f}"],
            ["", "", "", "Total:", f"${getattr(order, 'total_amount', 0):.2f}"]
        ])
        
        items_table = Table(items_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d5c8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            # Content style
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Product ID
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Product Title
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),  # Quantity, Price, Subtotal
            # Grid style
            ('GRID', (0, 0), (-1, -5), 1, colors.grey),
            # Totals style
            ('LINEABOVE', (-2, -3), (-1, -3), 1, colors.black),
            ('FONTNAME', (-2, -3), (-1, -1), 'Helvetica-Bold'),
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(items_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        logger.error(f"Error generating PDF for order {order_id}: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': 'Failed to generate PDF',
            'details': str(e)
        }, status=500)
